# -*- coding: utf-8 -*-
"""
SHAMEL Attendance System — views.py
Auto-reconstructed from helper scripts.
"""

# ── Standard library ────────────────────────────────────────────────────────
import os
import io
import re
import csv
import json
import logging
import threading
import base64
import tempfile
from collections import defaultdict
from datetime import datetime, timedelta

# ── Django core ──────────────────────────────────────────────────────────────
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import login as auth_login, logout as auth_logout, authenticate
from django.contrib.auth.decorators import login_required
from django.contrib.admin.views.decorators import staff_member_required
from django.contrib.auth.models import User
from django.contrib import messages
from django.http import (
    HttpResponse, HttpResponseBadRequest, JsonResponse, StreamingHttpResponse, Http404
)
from django.urls import reverse
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_POST
from django.utils import timezone
from django.db import models
from django.db.models import Count, Q, Avg

# ── Project models ───────────────────────────────────────────────────────────
from .models import (
    College, Department, Classroom, Course, Teacher, Student,
    Coordinator, Schedule, Enrollment, LectureSession,
    AIAttendanceLog, StudentFaceEmbedding, TeacherFaceEmbedding,
    CameraSource, GateLog, Notification, SupportTicket,
    FinancialStatus, AuditLog, MedicalExcuse,
    Exam, ExamSeat, CourseEvaluation, SystemConfig, AsyncTask,
)

# ── Face engine (InsightFace — single active engine) ─────────────────────────
from . import face_engine as _fe

# face_recognition (dlib) is disabled — InsightFace is the only engine.
face_recognition = None
FACE_RECOGNITION_AVAILABLE = False

try:
    import cv2
    CV2_AVAILABLE = True
except ImportError:
    cv2 = None
    CV2_AVAILABLE = False

try:
    import numpy as np
    NUMPY_AVAILABLE = True
except ImportError:
    np = None
    NUMPY_AVAILABLE = False

FACE_ENGINE_AVAILABLE = _fe.available()

logger = logging.getLogger(__name__)

# University attendance policy — change here to update all reports/dashboard/exports
ATTENDANCE_PASS_THRESHOLD = 75  # percent

# ── Semester helpers ─────────────────────────────────────────────────────────
ARABIC_ORDINALS = ['الأول','الثاني','الثالث','الرابع','الخامس',
                   'السادس','السابع','الثامن','التاسع','العاشر',
                   'الحادي عشر','الثاني عشر']
ARABIC_YEAR_ORDINALS = ['الأولى','الثانية','الثالثة','الرابعة',
                         'الخامسة','السادسة']

def get_semester_choices(num_semesters=8):
    """Return list of (value, label) for num_semesters semesters, with year context."""
    choices = []
    for i in range(1, num_semesters + 1):
        year_num = (i - 1) // 2
        sem_in_year = ((i - 1) % 2) + 1
        sem_ord = ARABIC_ORDINALS[i - 1] if i <= len(ARABIC_ORDINALS) else str(i)
        year_ord = ARABIC_YEAR_ORDINALS[year_num] if year_num < len(ARABIC_YEAR_ORDINALS) else f'{year_num+1}'
        sem_label = f'الفصل {sem_ord} — السنة {year_ord}'
        choices.append((str(i), sem_label))
    return choices

# Standard semester sets by college type (value, label)
SEMESTER_CHOICES_4Y  = get_semester_choices(8)   # 4-year programs (Law, Science, Commerce, etc.)
SEMESTER_CHOICES_5Y  = get_semester_choices(10)  # 5-year programs (Engineering, Pharmacy)
SEMESTER_CHOICES_6Y  = get_semester_choices(12)  # 6-year programs (Medicine, Dentistry)

# ── In-memory face cache ─────────────────────────────────────────────────────
known_face_encodings = []
known_face_names = []
_face_lock = threading.Lock()

# ── Camera stop-flag registry ─────────────────────────────────────────────────
# Maps camera_index → threading.Event.  gen_frames() checks this every frame.
# stop_camera_stream() sets the event → gen_frames exits cleanly → cap.release().
_camera_stop_flags: dict = {}
_camera_flags_lock = threading.Lock()

# ── Stub models that may not exist in all migrations ────────────────────────
# ClassroomPermission and GateEntryLog are referenced in helper scripts.
# Provide safe stubs so the module imports without error.
try:
    from .models import ClassroomPermission
except ImportError:
    class ClassroomPermission:
        objects = type('M', (), {'filter': staticmethod(lambda **kw: []),
                                  'all': staticmethod(lambda: [])})()

try:
    from .models import GateEntryLog
except ImportError:
    class GateEntryLog:
        objects = type('M', (), {
            'all': staticmethod(lambda: _EmptyQS()),
            'filter': staticmethod(lambda **kw: _EmptyQS()),
        })()

class _EmptyQS:
    def filter(self, **kw): return self
    def order_by(self, *a): return self
    def count(self): return 0
    def values(self, *a): return self
    def annotate(self, **kw): return []
    def __iter__(self): return iter([])


# ═══════════════════════════════════════════════════════════════════════════════
# HELPER FUNCTIONS
# ═══════════════════════════════════════════════════════════════════════════════

def _unique_username(*candidates):
    """Build a unique, valid username from the first non-empty candidate
    (email local-part, student code, or name). Registration forms collect
    name/email/password but no explicit username, so we derive one."""
    base = ''
    for c in candidates:
        c = (c or '').strip()
        if c:
            base = c.split('@')[0] if '@' in c else c
            break
    base = re.sub(r'[^a-zA-Z0-9_.]', '', base.replace(' ', '_')) or 'user'
    username, n = base, 1
    while User.objects.filter(username=username).exists():
        n += 1
        username = f'{base}{n}'
    return username


def _redirect_by_role(request):
    """Redirect the logged-in user to their role-appropriate dashboard."""
    user = request.user
    if not user.is_authenticated:
        return redirect('login')
    # gate_staff check first — before is_staff to avoid mis-routing
    if user.groups.filter(name='gate_staff').exists():
        return redirect('gate_page')
    if user.is_superuser or user.is_staff:
        return redirect('admin_panel')
    if Coordinator.objects.filter(auth_user=user).exists():
        return redirect('coordinator_dashboard')
    if Teacher.objects.filter(auth_user=user).exists():
        return redirect('professor_dashboard')
    if Student.objects.filter(auth_user=user).exists():
        return redirect('student_dashboard')
    return redirect('login')


def log_audit(request, action, target_model='', target_id='', description=''):
    """Write an AuditLog entry (best-effort, never raises)."""
    try:
        ip = (request.META.get('HTTP_X_FORWARDED_FOR', '') or
              request.META.get('REMOTE_ADDR', ''))
        ip = ip.split(',')[0].strip() or None
        AuditLog.objects.create(
            user=request.user if request.user.is_authenticated else None,
            action=action,
            target_model=target_model,
            target_id=str(target_id),
            description=description,
            ip_address=ip,
        )
    except Exception:
        pass


def load_known_faces():
    """Load face embeddings into in-memory cache via InsightFace engine.

    Only loads is_allowed_entry=True students + all teachers.
    Skips embeddings whose dimension doesn't match the active engine.
    """
    global known_face_encodings, known_face_names
    if not NUMPY_AVAILABLE:
        return
    expected_dim = _fe.embedding_dim()
    with _face_lock:
        known_face_encodings.clear()
        known_face_names.clear()
        for emb in (StudentFaceEmbedding.objects
                    .select_related('student')
                    .filter(student__is_allowed_entry=True)
                    .only('embedding', 'extra_embeddings', 'student__name')):
            try:
                vec = emb.embedding if isinstance(emb.embedding, list) else list(emb.embedding)
                if len(vec) == expected_dim:
                    known_face_encodings.append(np.array(vec, dtype=np.float32))
                    known_face_names.append(emb.student.name)
                for extra in (emb.extra_embeddings or []):
                    if isinstance(extra, list) and len(extra) == expected_dim:
                        known_face_encodings.append(np.array(extra, dtype=np.float32))
                        known_face_names.append(emb.student.name)
            except Exception:
                pass
        for emb in (TeacherFaceEmbedding.objects
                    .select_related('teacher')
                    .only('face_vector', 'extra_embeddings', 'teacher__name')):
            try:
                vec = emb.face_vector if isinstance(emb.face_vector, list) else list(emb.face_vector)
                if len(vec) == expected_dim:
                    known_face_encodings.append(np.array(vec, dtype=np.float32))
                    known_face_names.append(emb.teacher.name)
                for extra in (emb.extra_embeddings or []):
                    if isinstance(extra, list) and len(extra) == expected_dim:
                        known_face_encodings.append(np.array(extra, dtype=np.float32))
                        known_face_names.append(emb.teacher.name)
            except Exception:
                pass


def match_face_from_db(live_encoding: list) -> tuple[str | None, str | None, int | None]:
    """On-demand DB match using InsightFace engine — searches ALL batches for best match.

    Returns (name, person_type, pk) or (None, None, None).
    Skips embeddings whose dimension doesn't match the active engine.
    """
    if not NUMPY_AVAILABLE or live_encoding is None:
        return None, None, None
    probe = np.array(live_encoding, dtype=np.float32)
    expected_dim = _fe.embedding_dim()
    CHUNK = 200

    def _best_in_collection(queryset, vec_field, extra_field, name_getter, pk_getter, person_type):
        """Scan all batches (primary + extra_embeddings) and return best match."""
        best_score = -1.0
        best_name = None
        best_pk = None
        offset = 0
        while True:
            batch = list(queryset[offset: offset + CHUNK])
            if not batch:
                break
            for emb in batch:
                try:
                    # Collect all vectors: primary + any extra angles
                    all_vecs = []
                    raw = getattr(emb, vec_field)
                    primary = raw if isinstance(raw, list) else list(raw)
                    if len(primary) == expected_dim:
                        all_vecs.append(primary)
                    for extra in (getattr(emb, extra_field, None) or []):
                        if isinstance(extra, list) and len(extra) == expected_dim:
                            all_vecs.append(extra)
                    if not all_vecs:
                        continue
                    idx, score = _fe.match(all_vecs, probe.tolist())
                    if idx >= 0 and score > best_score:
                        best_score = score
                        best_name = name_getter(emb)
                        best_pk = pk_getter(emb)
                except Exception:
                    pass
            offset += CHUNK
        if best_name:
            return best_name, person_type, best_pk
        return None, None, None

    s_name, s_type, s_pk = _best_in_collection(
        StudentFaceEmbedding.objects.select_related('student')
            .only('embedding', 'extra_embeddings', 'student__id', 'student__name'),
        'embedding', 'extra_embeddings',
        lambda e: e.student.name,
        lambda e: e.student.pk,
        'student',
    )
    t_name, t_type, t_pk = _best_in_collection(
        TeacherFaceEmbedding.objects.select_related('teacher')
            .only('face_vector', 'extra_embeddings', 'teacher__id', 'teacher__name'),
        'face_vector', 'extra_embeddings',
        lambda e: e.teacher.name,
        lambda e: e.teacher.pk,
        'teacher',
    )
    # Return whichever produced a match; students take precedence on tie
    if s_name:
        return s_name, s_type, s_pk
    return t_name, t_type, t_pk


def _find_best_camera(preferred_index=0):
    """Return the best available camera index.
    If preferred_index fails, try external cameras (1,2) then fallback to 0."""
    if not CV2_AVAILABLE:
        return preferred_index
    for idx in [preferred_index, 1, 2, 0]:
        cap = cv2.VideoCapture(idx, cv2.CAP_DSHOW)
        if cap.isOpened():
            ret, _ = cap.read()
            cap.release()
            if ret:
                return idx
    return preferred_index


def gen_frames(camera_index=0):
    """Yield MJPEG frames from camera with optional face detection overlay."""
    if not CV2_AVAILABLE:
        return
    actual_index = _find_best_camera(camera_index)

    # Register a stop-event for this camera index.
    # Any previous generator for this camera is signalled to stop.
    with _camera_flags_lock:
        old_event = _camera_stop_flags.get(camera_index)
        if old_event:
            old_event.set()  # stop previous stream for same camera
        stop_event = threading.Event()
        _camera_stop_flags[camera_index] = stop_event

    cap = cv2.VideoCapture(actual_index, cv2.CAP_DSHOW)
    cap.set(cv2.CAP_PROP_BUFFERSIZE, 1)  # minimal buffer = faster stop response

    _fail_count   = 0
    _MAX_FAILS    = 30
    _frame_counter = 0
    _FACE_EVERY_N  = 5
    _last_overlays = []
    try:
        while not stop_event.is_set():
            success, frame = cap.read()
            if not success:
                _fail_count += 1
                if _fail_count >= _MAX_FAILS:
                    break
                continue
            _fail_count = 0
            _frame_counter += 1

            # Check stop flag AFTER read so we exit immediately next iteration
            if stop_event.is_set():
                break

            # Face detection — every _FACE_EVERY_N frames
            if FACE_ENGINE_AVAILABLE and CV2_AVAILABLE and NUMPY_AVAILABLE and (_frame_counter % _FACE_EVERY_N == 0):
                try:
                    rgb_frame = frame[:, :, ::-1]
                    app = _fe._get_insightface()
                    new_overlays = []
                    if app is not None:
                        faces = app.get(rgb_frame)
                        for face in faces:
                            x1, y1, x2, y2 = [int(v) for v in face.bbox]
                            probe = [float(x) for x in face.normed_embedding]
                            name = 'غير معروف'
                            if known_face_encodings:
                                idx, score = _fe.match(
                                    [e.tolist() for e in known_face_encodings], probe
                                )
                                if idx >= 0:
                                    name = known_face_names[idx]
                            color = (0, 0, 220) if name == 'غير معروف' else (0, 220, 0)
                            new_overlays.append((y1, x2, y2, x1, color, name))
                    _last_overlays = new_overlays
                except Exception:
                    pass

            for (top, right, bottom, left, color, label) in _last_overlays:
                cv2.rectangle(frame, (left, top), (right, bottom), color, 2)
                cv2.putText(frame, label, (left, top - 10),
                            cv2.FONT_HERSHEY_SIMPLEX, 0.55, color, 2)

            ret, buffer = cv2.imencode('.jpg', frame)
            if not ret:
                continue
            yield (b'--frame\r\nContent-Type: image/jpeg\r\n\r\n' +
                   buffer.tobytes() + b'\r\n')
    except GeneratorExit:
        pass
    except Exception:
        pass
    finally:
        stop_event.set()   # mark as stopped in case something else checks
        cap.release()
        cap.release()      # call twice — some CAP_DSHOW handles need double-release
        del cap            # force Python GC release of DirectShow handle
        with _camera_flags_lock:
            if _camera_stop_flags.get(camera_index) is stop_event:
                del _camera_stop_flags[camera_index]


# ═══════════════════════════════════════════════════════════════════════════════
# AUTH VIEWS
# ═══════════════════════════════════════════════════════════════════════════════

from django.views.decorators.csrf import ensure_csrf_cookie


def _sync_display_name(user):
    """On login, copy the real Arabic name from profile → user.first_name
    so base.html's get_full_name shows the actual name, not 'tchr_13'."""
    if user.first_name:
        return  # already has a name set
    try:
        t = Teacher.objects.filter(auth_user=user).first()
        if t and t.name:
            parts = t.name.strip().split()
            user.first_name = parts[0] if parts else t.name
            user.last_name  = ' '.join(parts[1:]) if len(parts) > 1 else ''
            user.save(update_fields=['first_name', 'last_name'])
            return
    except Exception:
        pass
    try:
        s = Student.objects.filter(auth_user=user).first()
        if s and s.name:
            parts = s.name.strip().split()
            user.first_name = parts[0] if parts else s.name
            user.last_name  = ' '.join(parts[1:]) if len(parts) > 1 else ''
            user.save(update_fields=['first_name', 'last_name'])
            return
    except Exception:
        pass
    try:
        c = Coordinator.objects.filter(auth_user=user).first()
        if c and c.name:
            parts = c.name.strip().split()
            user.first_name = parts[0] if parts else c.name
            user.last_name  = ' '.join(parts[1:]) if len(parts) > 1 else ''
            user.save(update_fields=['first_name', 'last_name'])
    except Exception:
        pass


@ensure_csrf_cookie  # always plant a fresh csrftoken cookie on the login page
def login_view(request):
    if request.user.is_authenticated:
        return _redirect_by_role(request)
    if request.method == 'POST':
        username = request.POST.get('username', '').strip()
        password = request.POST.get('password', '')
        user = authenticate(request, username=username, password=password)
        if user:
            auth_login(request, user)
            # Sync full name from profile so base.html shows real name not username
            _sync_display_name(user)
            log_audit(request, 'LOGIN', 'User', user.pk, f'{username} logged in')
            next_url = request.GET.get('next', '')
            if next_url:
                return redirect(next_url)
            return _redirect_by_role(request)
        resp = render(request, 'attendance/university_login.html',
                      {'error': 'اسم المستخدم أو كلمة المرور غير صحيحة — حاول مجدداً.'})
        resp['Cache-Control'] = 'no-store, no-cache, must-revalidate'
        return resp
    resp = render(request, 'attendance/university_login.html')
    # Prevent browser/PWA from caching the login page — a cached page
    # carries a stale CSRF token which causes "403 CSRF token incorrect".
    resp['Cache-Control'] = 'no-store, no-cache, must-revalidate'
    resp['Pragma'] = 'no-cache'
    return resp


def logout_view(request):
    log_audit(request, 'LOGOUT', 'User', request.user.pk if request.user.is_authenticated else '',
              'User logged out')
    auth_logout(request)
    return redirect('login')


# ═══════════════════════════════════════════════════════════════════════════════
# CAMERA / SCAN VIEWS
# ═══════════════════════════════════════════════════════════════════════════════

@login_required
def scan_page(request):
    # Block coordinators — scan station is admin/staff only
    if not (request.user.is_staff or request.user.is_superuser):
        if Coordinator.objects.filter(auth_user=request.user).exists():
            return _redirect_by_role(request)
    camera_sources = CameraSource.objects.filter(is_active=True, is_gate=False)
    sessions = LectureSession.objects.filter(is_active=True).select_related(
        'schedule__course', 'schedule__teacher', 'schedule__classroom'
    )
    return render(request, 'attendance/scan.html', {
        'cameras': camera_sources,
        'camera_sources': camera_sources,  # template uses camera_sources
        'active_sessions': sessions,
    })


@login_required
def video_feed(request):
    # Accept both 'camera' and 'source' params for compatibility
    cam_param = request.GET.get('camera') or request.GET.get('source', '0')
    try:
        camera_index = int(cam_param)
    except (ValueError, TypeError):
        camera_index = 0
    if not CV2_AVAILABLE:
        return HttpResponse(status=503)
    try:
        response = StreamingHttpResponse(
            gen_frames(camera_index),
            content_type='multipart/x-mixed-replace; boundary=frame'
        )
        # Tell browser not to cache the stream
        response['Cache-Control'] = 'no-cache, no-store'
        return response
    except Exception:
        return HttpResponse(status=503)


@login_required
@csrf_exempt
def stop_camera(request):
    """Called via navigator.sendBeacon() when the user leaves the scan page.
    Sets the stop-event for the requested camera so gen_frames exits immediately."""
    cam_param = request.POST.get('camera', '0') or request.GET.get('camera', '0')
    try:
        camera_index = int(cam_param)
    except (ValueError, TypeError):
        camera_index = 0
    with _camera_flags_lock:
        event = _camera_stop_flags.get(camera_index)
        if event:
            event.set()
    return HttpResponse(status=204)


@login_required
def check_status(request):
    active_sessions = LectureSession.objects.filter(is_active=True).count()
    recent = AIAttendanceLog.objects.order_by('-timestamp').first()
    return JsonResponse({
        'active_sessions': active_sessions,
        'last_scan': recent.timestamp.isoformat() if recent else None,
        'face_recognition': FACE_ENGINE_AVAILABLE,
    })


@login_required
def recent_scans(request):
    logs = AIAttendanceLog.objects.select_related('student', 'schedule__course') \
                                   .order_by('-timestamp')[:20]
    data = []
    for log in logs:
        data.append({
            'student': log.student.name if log.student else '',
            'course': log.schedule.course.title if (log.schedule and log.schedule.course) else '',
            'status': log.status,
            'timestamp': log.timestamp.strftime('%Y-%m-%d %H:%M:%S'),
            'confidence': log.confidence_score,
        })
    return JsonResponse({'scans': data})


@login_required
def live_stats(request):
    today = timezone.now().date()
    total_today = AIAttendanceLog.objects.filter(timestamp__date=today).count()
    present_today = AIAttendanceLog.objects.filter(timestamp__date=today, status='Present').count()
    active_sessions = LectureSession.objects.filter(is_active=True).count()
    return JsonResponse({
        'total_today': total_today,
        'present_today': present_today,
        'active_sessions': active_sessions,
    })


@login_required
def attendance_logs(request):
    logs = AIAttendanceLog.objects.select_related(
        'student', 'schedule__course', 'schedule__teacher'
    ).order_by('-timestamp')

    # Coordinator: restrict to their college only
    coordinator = Coordinator.objects.filter(auth_user=request.user).first()
    if coordinator and not (request.user.is_staff or request.user.is_superuser):
        logs = logs.filter(schedule__course__college=coordinator.college)

    date_f = request.GET.get('date', '')
    status_f = request.GET.get('status', '')
    student_q = request.GET.get('student', '')

    if date_f:
        logs = logs.filter(timestamp__date=date_f)
    if status_f:
        logs = logs.filter(status=status_f)
    if student_q:
        logs = logs.filter(student__name__icontains=student_q)

    return render(request, 'attendance/attendance_logs.html', {
        'logs': logs[:200],
        'filters': {'date': date_f, 'status': status_f, 'student': student_q},
    })


def attendance_success(request):
    return render(request, 'attendance/attendance_success.html')


def attendance_error(request):
    return render(request, 'attendance/attendance_error.html')


# ═══════════════════════════════════════════════════════════════════════════════
# ADMIN VIEWS
# ═══════════════════════════════════════════════════════════════════════════════

@login_required
@staff_member_required
def admin_control_panel(request):
    students_count = Student.objects.count()
    teachers_count = Teacher.objects.count()
    courses_count  = Course.objects.count()
    today          = timezone.now().date()
    attendance_today  = AIAttendanceLog.objects.filter(timestamp__date=today).count()
    active_sessions   = LectureSession.objects.filter(is_active=True).count()
    recent_logs       = AIAttendanceLog.objects.select_related('student', 'schedule__course') \
                                               .order_by('-timestamp')[:10]
    unread_notifs = Notification.objects.filter(user=request.user, is_read=False).count()
    open_tickets  = SupportTicket.objects.filter(status='open').count()

    # Context vars expected by admin_panel.html template
    teachers   = Teacher.objects.select_related('department', 'college').order_by('name')[:10]
    classrooms = Classroom.objects.order_by('name')[:12]

    # training_progress: % of students that have a face embedding enrolled
    total_s = students_count or 1
    enrolled_faces = StudentFaceEmbedding.objects.count()
    training_progress = min(round(enrolled_faces / total_s * 100), 100)

    setup_done = SystemConfig.objects.filter(key='setup_done').exists()

    return render(request, 'attendance/admin_panel.html', {
        # legacy names kept for backward-compat
        'students_count':   students_count,
        'teachers_count':   teachers_count,
        'courses_count':    courses_count,
        'attendance_today': attendance_today,
        'active_sessions':  active_sessions,
        'recent_logs':      recent_logs,
        'unread_notifs':    unread_notifs,
        'open_tickets':     open_tickets,
        # names expected by template
        'total_students':    students_count,
        'active_faculty':    teachers_count,
        'training_progress': training_progress,
        'teachers':          teachers,
        'classrooms':        classrooms,
        'setup_done':        setup_done,
    })


@login_required
def faculty_management(request):
    is_admin = request.user.is_superuser or request.user.is_staff
    coordinator = Coordinator.objects.filter(auth_user=request.user).first()
    if not is_admin and not coordinator:
        return _redirect_by_role(request)

    teachers = Teacher.objects.select_related('department', 'college').order_by('name')

    # Coordinators can only see their own college
    if coordinator and not is_admin:
        teachers = teachers.filter(college=coordinator.college)

    college_f = request.GET.get('college_id', '')
    dept_f = request.GET.get('department_id', '')
    q = request.GET.get('q', '')
    if college_f and is_admin:
        teachers = teachers.filter(college_id=college_f)
    if dept_f:
        teachers = teachers.filter(department_id=dept_f)
    if q:
        teachers = teachers.filter(name__icontains=q)

    # Limit college/department filter options by role
    if coordinator and not is_admin:
        colleges    = College.objects.filter(pk=coordinator.college_id)
        departments = Department.objects.filter(college=coordinator.college)
    else:
        colleges    = College.objects.all()
        departments = Department.objects.all()

    return render(request, 'attendance/faculty_management.html', {
        'teachers':    teachers,
        'colleges':    colleges,
        'departments': departments,
        'coordinator': coordinator,
        'is_admin':    is_admin,
        'filters': {'college_id': college_f, 'department_id': dept_f, 'q': q},
    })


@login_required
@staff_member_required
def reports_view(request):
    from django.db.models import Avg, FloatField
    from django.db.models.functions import ExtractWeekDay
    today = timezone.now().date()
    week_ago = today - timedelta(days=7)

    # ── Weekly attendance data (Sun–Fri) ──────────────────────────────────────
    day_map = {1: 'Sun', 2: 'Mon', 3: 'Tue', 4: 'Wed', 5: 'Thu', 6: 'Fri', 7: 'Sat'}
    weekly_qs = (
        AIAttendanceLog.objects
        .filter(timestamp__date__gte=week_ago, status='Present')
        .annotate(dow=ExtractWeekDay('timestamp'))
        .values('dow')
        .annotate(count=Count('id'))
    )
    weekly_data = {'Sun': 0, 'Mon': 0, 'Tue': 0, 'Wed': 0, 'Thu': 0, 'Fri': 0}
    for row in weekly_qs:
        label = day_map.get(row['dow'], '')
        if label in weekly_data:
            weekly_data[label] = row['count']

    # ── Total sessions ────────────────────────────────────────────────────────
    total_sessions = LectureSession.objects.count()

    # ── Avg attendance % ─────────────────────────────────────────────────────
    total_logs = AIAttendanceLog.objects.count()
    present_logs = AIAttendanceLog.objects.filter(status='Present').count()
    avg_attendance = round((present_logs / total_logs * 100), 1) if total_logs else 0

    # ── AI Engine accuracy (avg confidence of present scans) ──────────────────
    avg_conf = (
        AIAttendanceLog.objects
        .filter(status='Present', confidence_score__isnull=False)
        .aggregate(avg=Avg('confidence_score', output_field=FloatField()))
        ['avg'] or 0
    )
    ai_accuracy = round(avg_conf * 100, 1)

    # ── Top departments by attendance ─────────────────────────────────────────
    top_departments = (
        AIAttendanceLog.objects
        .filter(status='Present', student__department__isnull=False)
        .values('student__department__name')
        .annotate(count=Count('id'))
        .order_by('-count')[:5]
    )
    top_depts = [
        {'name': r['student__department__name'], 'count': r['count']}
        for r in top_departments
    ]

    # ── Daily counts (for any other chart) ───────────────────────────────────
    daily_counts = (AIAttendanceLog.objects
                    .filter(timestamp__date__gte=week_ago)
                    .values('timestamp__date')
                    .annotate(count=Count('id'))
                    .order_by('timestamp__date'))

    return render(request, 'attendance/reports.html', {
        'daily_counts':    list(daily_counts),
        'total_students':  Student.objects.count(),
        'total_teachers':  Teacher.objects.count(),
        'total_sessions':  total_sessions,
        'avg_attendance':  avg_attendance,
        'ai_accuracy':     ai_accuracy,
        'weekly_data':     weekly_data,
        'top_departments': top_depts,
    })


@login_required
def stop_session(request, session_id):
    session = get_object_or_404(LectureSession, pk=session_id)
    # Teacher can only stop their own session; admin can stop any
    if not request.user.is_staff:
        teacher = get_object_or_404(Teacher, auth_user=request.user)
        if session.schedule.teacher != teacher:
            messages.error(request, 'غير مصرح لك بإيقاف هذه الجلسة.')
            return redirect('professor_dashboard')
    session.is_active = False
    session.actual_end_time = timezone.now()
    session.save()
    log_audit(request, 'STOP_SESSION', 'LectureSession', session_id)
    messages.success(request, 'تم إيقاف الجلسة.')
    if Teacher.objects.filter(auth_user=request.user).exists():
        return redirect('professor_dashboard')
    return redirect('admin_panel')


@login_required
def get_chancellor_stats(request):
    today = timezone.now().date()
    total_students  = Student.objects.count()
    trained_students = Student.objects.filter(is_trained=True).count()
    training_progress = round(trained_students / total_students * 100) if total_students else 0
    total_logs = AIAttendanceLog.objects.count()
    correct_logs = AIAttendanceLog.objects.filter(status='Present').count()
    global_accuracy = round(correct_logs / total_logs * 100) if total_logs else 0
    return JsonResponse({
        'students': total_students,
        'total_students': total_students,
        'teachers': Teacher.objects.count(),
        'courses': Course.objects.count(),
        'attendance_today': AIAttendanceLog.objects.filter(timestamp__date=today).count(),
        'active_sessions': LectureSession.objects.filter(is_active=True).count(),
        'colleges': College.objects.count(),
        'global_accuracy': global_accuracy,
        'training_progress': training_progress,
    })


@login_required
@staff_member_required
def export_teachers_csv(request):
    teachers = Teacher.objects.select_related('department', 'college').order_by('name')
    response = HttpResponse(content_type='text/csv; charset=utf-8-sig')
    from datetime import date
    response['Content-Disposition'] = f'attachment; filename="SHAMEL_Teachers_Export_{date.today()}.csv"'
    writer = csv.writer(response)
    writer.writerow(['ID', 'Name', 'Degree', 'Major', 'Department', 'College', 'Email', 'Phone', 'Allowed Entry'])
    for t in teachers:
        writer.writerow([
            t.teacher_id, t.name, t.academic_degree, t.major,
            t.department.name if t.department else '',
            t.college.college_name if t.college else '',
            t.university_email, t.phone_number, t.is_allowed_entry,
        ])
    return response


@login_required
def upload_face(request, user_type, user_id):
    """Upload a face image and store embedding."""
    # Only staff/admin may manage face embeddings for arbitrary users.
    if not (request.user.is_staff or request.user.is_superuser):
        messages.error(request, 'غير مصرح.')
        return _redirect_by_role(request)

    # Load the person for context
    person = None
    if user_type == 'teacher':
        person = get_object_or_404(Teacher, pk=user_id)
    elif user_type == 'student':
        person = get_object_or_404(Student, pk=user_id)

    if request.method == 'GET':
        return render(request, 'attendance/upload_face.html', {
            'person': person,
            'user_type': user_type,
            'user_id': user_id,
        })

    if request.method == 'POST' and request.FILES.get('face_image'):
        img_file = request.FILES['face_image']
        face_enrolled = False
        if user_type == 'student':
            obj = get_object_or_404(Student, pk=user_id)
            obj.face_image = img_file
            obj.save(update_fields=['face_image'])
            if FACE_ENGINE_AVAILABLE and NUMPY_AVAILABLE:
                try:
                    import cv2 as _cv2
                    raw = np.frombuffer(open(obj.face_image.path, 'rb').read(), dtype=np.uint8)
                    img = _cv2.imdecode(raw, _cv2.IMREAD_COLOR)
                    if img is not None:
                        enc = _fe.encode(img[:, :, ::-1])  # BGR→RGB
                        if enc:
                            StudentFaceEmbedding.objects.update_or_create(
                                student=obj, defaults={'embedding': enc}
                            )
                            load_known_faces()
                            face_enrolled = True
                except Exception as e:
                    logger.warning('Face encoding error: %s', e)
        elif user_type == 'teacher':
            obj = get_object_or_404(Teacher, pk=user_id)
            obj.face_image = img_file
            obj.save(update_fields=['face_image'])
            if FACE_ENGINE_AVAILABLE and NUMPY_AVAILABLE:
                try:
                    import cv2 as _cv2
                    raw = np.frombuffer(open(obj.face_image.path, 'rb').read(), dtype=np.uint8)
                    img = _cv2.imdecode(raw, _cv2.IMREAD_COLOR)
                    if img is not None:
                        enc = _fe.encode(img[:, :, ::-1])
                        if enc:
                            TeacherFaceEmbedding.objects.update_or_create(
                                teacher=obj, defaults={'face_vector': enc}
                            )
                            load_known_faces()
                            face_enrolled = True
                except Exception as e:
                    logger.warning('Face encoding error: %s', e)
        if face_enrolled:
            messages.success(request, 'تم رفع الصورة وتسجيل بصمة الوجه بنجاح.')
        else:
            messages.warning(request, 'تم رفع الصورة لكن لم يتم اكتشاف وجه — تأكد أن الوجه واضح في الصورة وأعد الرفع.')
    return _redirect_by_role(request)


@login_required
@require_POST
def toggle_user_access(request, user_type, user_id):
    """
    Gate-entry toggle.
    STUDENTS only: toggled based on registration status / fees / disciplinary.
    TEACHERS: always allowed entry — they are employed staff, not enrollees.
              Do NOT toggle teacher access; return 400 if attempted.
    Only admin/staff or coordinators scoped to the student's college may toggle.
    """
    is_admin = request.user.is_staff or request.user.is_superuser
    coord = Coordinator.objects.filter(auth_user=request.user).first()
    if not is_admin and not coord:
        messages.error(request, 'غير مصرح لك بتغيير صلاحيات الدخول.')
        return redirect(request.META.get('HTTP_REFERER', '/'))
    if user_type == 'student':
        obj = get_object_or_404(Student, pk=user_id)
        obj.is_allowed_entry = not obj.is_allowed_entry
        obj.save(update_fields=['is_allowed_entry'])
        # Email student if they just became ineligible
        if not obj.is_allowed_entry:
            try:
                from .email_utils import notify_student_ineligible
                notify_student_ineligible(obj)
            except Exception:
                pass
        log_audit(request, 'TOGGLE_STUDENT_ACCESS', user_type, user_id)
    elif user_type == 'teacher':
        # Teachers are always permitted entry — access toggle does not apply.
        messages.warning(request,
            'الأساتذة لا يخضعون لنظام التصاريح. الدخول مضمون دائماً بحكم عقد التوظيف.')
    else:
        messages.error(request, 'نوع مستخدم غير معروف.')
    return redirect(request.META.get('HTTP_REFERER', '/'))


# ── Gate ─────────────────────────────────────────────────────────────────────

# Cooldown: don't re-log same person within this many seconds
_GATE_COOLDOWN_SEC = 30
_gate_cooldown: dict[str, float] = {}  # person_key → last_log_epoch
_gate_cooldown_lock = threading.Lock()


@login_required
def gate_page(request):
    user = request.user
    is_gate_staff = user.groups.filter(name__in=['gate_staff', 'GATE_STAFF']).exists()
    if not (user.is_staff or user.is_superuser or is_gate_staff):
        from django.http import HttpResponseForbidden
        return HttpResponseForbidden('Access denied.')
    today = timezone.now().date()
    recent_logs = GateLog.objects.order_by('-timestamp').select_related('student', 'teacher')[:50]
    today_count   = GateLog.objects.filter(timestamp__date=today).count()
    allowed_count = GateLog.objects.filter(timestamp__date=today, status='Allowed').count()
    denied_count  = GateLog.objects.filter(timestamp__date=today, status='Denied').count()
    return render(request, 'attendance/gate.html', {
        'recent_logs':   recent_logs,
        'today_count':   today_count,
        'allowed_count': allowed_count,
        'denied_count':  denied_count,
    })


@login_required
@require_POST
@csrf_exempt
def gate_scan_api(request):
    """Receive webcam frame, match face, log GateLog, return person info."""
    if not FACE_ENGINE_AVAILABLE or not CV2_AVAILABLE or not NUMPY_AVAILABLE:
        return JsonResponse({'ok': False, 'error': 'engine_unavailable'})

    b64 = request.POST.get('frame', '')
    if not b64:
        return JsonResponse({'ok': False, 'error': 'no_frame'})

    try:
        if ',' in b64:
            b64 = b64.split(',')[1]
        img_bytes = base64.b64decode(b64)
        img_arr = cv2.imdecode(np.frombuffer(img_bytes, dtype=np.uint8), cv2.IMREAD_COLOR)
        if img_arr is None:
            return JsonResponse({'ok': False, 'error': 'bad_image'})

        import time as _time

        # Encode ALL faces in one pass (multi-face support)
        all_faces = _fe.encode_all(img_arr[:, :, ::-1])

        if not all_faces:
            return JsonResponse({
                'ok': True, 'detected': False, 'results': [],
                'width': img_arr.shape[1], 'height': img_arr.shape[0],
            })

        now_ts   = _time.time()
        results  = []

        for face_data in all_faces:
            probe = face_data['embedding']
            bbox  = face_data['bbox']

            matched_name, matched_type, matched_pk = match_face_from_db(probe)
            if not matched_name:
                results.append({'bbox': bbox, 'matched': False})
                continue

            # Use PK for exact lookup — icontains on name can match the wrong person
            student = Student.objects.filter(pk=matched_pk).first() if matched_type == 'student' else None
            teacher = Teacher.objects.filter(pk=matched_pk).first() if matched_type == 'teacher' else None

            is_allowed  = True
            deny_reason = ''
            if student:
                is_allowed  = bool(student.is_allowed_entry)
                deny_reason = 'غير مسموح بالدخول (تعليق إداري أو رسوم غير مسددة)' if not is_allowed else ''

            status_str = 'Allowed' if is_allowed else 'Denied'
            person_key = f'{matched_name}_{status_str}'

            with _gate_cooldown_lock:
                last = _gate_cooldown.get(person_key, 0)
                in_cooldown = (now_ts - last) < _GATE_COOLDOWN_SEC
                if not in_cooldown:
                    _gate_cooldown[person_key] = now_ts

            if not in_cooldown:
                GateLog.objects.create(
                    person_name=matched_name,
                    student=student,
                    teacher=teacher,
                    status=status_str,
                )

            person_data = {
                'name':     matched_name,
                'type':     'student' if student else 'teacher',
                'allowed':  is_allowed,
                'reason':   deny_reason,
                'cooldown': in_cooldown,
                'bbox':     bbox,
            }
            if student:
                person_data['student_code'] = student.student_code or ''
                person_data['phone']        = student.phone_number or ''
                person_data['registered']   = student.is_registered
                fees = FinancialStatus.objects.filter(student=student).first()
                person_data['fees_paid'] = fees.is_paid if fees else True
            elif teacher:
                person_data['phone'] = getattr(teacher, 'phone_number', '') or ''

            results.append({'matched': True, 'person': person_data, 'bbox': bbox})

        matched_results = [r for r in results if r.get('matched')]
        return JsonResponse({
            'ok': True,
            'detected': True,
            'matched': len(matched_results) > 0,
            'results': results,
            # Legacy single-person field for backward compat
            'person': matched_results[0]['person'] if matched_results else None,
            'width': img_arr.shape[1], 'height': img_arr.shape[0],
        })

    except Exception as e:
        logger.error('gate_scan_api error: %s', e, exc_info=True)
        return JsonResponse({'ok': False, 'error': str(e)})


# ── Schedule ─────────────────────────────────────────────────────────────────

@login_required
def schedule_view(request):
    coordinator = Coordinator.objects.filter(auth_user=request.user).first()
    is_admin    = request.user.is_staff or request.user.is_superuser
    teacher_obj = Teacher.objects.filter(auth_user=request.user).first()
    student_obj = Student.objects.filter(auth_user=request.user).first()

    # Role-based filtering
    if is_admin:
        schedules  = Schedule.objects.select_related('course', 'teacher', 'classroom', 'course__college', 'course__department').all()
        role       = 'admin'
    elif coordinator:
        # Coordinator sees only their college — isolation enforced here
        schedules  = Schedule.objects.filter(
            course__college=coordinator.college
        ).select_related('course', 'teacher', 'classroom', 'course__department')
        role       = 'coordinator'
    elif teacher_obj:
        # Teacher sees only schedules assigned to them by coordinator
        schedules  = Schedule.objects.filter(
            teacher=teacher_obj
        ).select_related('course', 'teacher', 'classroom', 'course__department')
        role       = 'teacher'
    elif student_obj:
        # Students redirected to their specialised view
        return redirect('student_schedule')
    else:
        schedules  = Schedule.objects.none()
        role       = 'none'

    # Filters (accessible to admin + coordinator)
    filter_day      = request.GET.get('day', '')
    filter_semester = request.GET.get('semester', '')
    filter_dept     = request.GET.get('dept', '')

    if filter_day:
        schedules = schedules.filter(day_of_week=filter_day)
    if filter_semester:
        schedules = schedules.filter(semester=filter_semester)
    if filter_dept and role in ('admin', 'coordinator'):
        schedules = schedules.filter(course__department_id=filter_dept)

    # Group by day
    DAYS_ORDER = ['Saturday', 'Sunday', 'Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday']
    DAY_AR = {
        'Saturday': 'السبت', 'Sunday': 'الأحد', 'Monday': 'الاثنين',
        'Tuesday': 'الثلاثاء', 'Wednesday': 'الأربعاء',
        'Thursday': 'الخميس', 'Friday': 'الجمعة',
    }
    schedules_list = list(schedules.order_by('day_of_week', 'start_time'))
    grouped = {}
    for s in schedules_list:
        day = s.day_of_week
        if day not in grouped:
            grouped[day] = {'ar': DAY_AR.get(day, day), 'schedules': []}
        grouped[day]['schedules'].append(s)
    grouped_schedules = {d: grouped[d] for d in DAYS_ORDER if d in grouped}

    # Departments for filter (coordinator: only their college depts)
    if coordinator and not is_admin:
        departments = Department.objects.filter(college=coordinator.college).order_by('name')
    else:
        departments = Department.objects.order_by('name')

    # Courses for add-schedule form
    if coordinator and not is_admin:
        courses = Course.objects.filter(college=coordinator.college).order_by('title')
    else:
        courses = Course.objects.order_by('title') if is_admin else Course.objects.none()

    # Teachers for add-schedule form
    if coordinator and not is_admin:
        teachers = Teacher.objects.filter(college=coordinator.college).order_by('name')
    else:
        teachers = Teacher.objects.order_by('name') if is_admin else Teacher.objects.none()

    classrooms = Classroom.objects.order_by('name')
    days       = DAYS_ORDER

    return render(request, 'attendance/schedule.html', {
        'grouped_schedules': grouped_schedules,
        'schedules':         schedules_list,
        'is_admin':          is_admin,
        'coordinator':       coordinator,
        'teacher_obj':       teacher_obj,
        'role':              role,
        'days':              days,
        'day_ar':            DAY_AR,
        'filter_day':        filter_day,
        'filter_semester':   filter_semester,
        'filter_dept':       filter_dept,
        'departments':       departments,
        'courses':           courses,
        'teachers':          teachers,
        'classrooms':        classrooms,
        'semesters':         SEMESTER_CHOICES_4Y,
    })


@login_required
def add_schedule(request):
    coordinator = Coordinator.objects.filter(auth_user=request.user).first()
    is_admin    = request.user.is_staff or request.user.is_superuser
    if not (is_admin or coordinator):
        messages.error(request, 'غير مصرح.')
        return redirect('schedule')

    if request.method == 'POST':
        try:
            from django.db import IntegrityError
            teacher_id   = request.POST.get('teacher_id', '').strip() or None
            classroom_id = request.POST.get('classroom_id', '').strip() or None
            course_id    = request.POST.get('course_id', '').strip()
            total_lec    = request.POST.get('total_lectures_required', '28').strip()

            # Basic validation — course is mandatory
            if not course_id:
                messages.error(request, 'يجب اختيار المادة الدراسية.')
                return redirect('add_schedule')

            sched = Schedule.objects.create(
                course_id               = course_id,
                teacher_id              = teacher_id,      # nullable: OK if not selected
                classroom_id            = classroom_id,    # nullable: OK if not selected
                day_of_week             = request.POST.get('day_of_week'),
                start_time              = request.POST.get('start_time'),
                end_time                = request.POST.get('end_time'),
                batch                   = request.POST.get('batch', ''),
                semester                = request.POST.get('semester', ''),
                total_lectures_required = int(total_lec) if total_lec.isdigit() else 28,
            )
            # Notify teacher about new assignment
            if teacher_id:
                try:
                    from .email_utils import notify_teacher_assignment
                    teacher_obj = Teacher.objects.get(pk=teacher_id)
                    notify_teacher_assignment(teacher_obj, sched)
                except Exception:
                    pass
            messages.success(request, 'تمت إضافة الحصة الدراسية بنجاح.')
        except IntegrityError as e:
            messages.error(request, 'تعذّر حفظ الحصة — تأكد من اكتمال جميع الحقول المطلوبة.')
        except Exception as e:
            messages.error(request, f'خطأ غير متوقع: {e}')
        return redirect('schedule')

    # GET — render form
    # Search/filter support for large datasets (1000+ courses)
    course_q = request.GET.get('course_q', '').strip()
    DAYS = ['Saturday', 'Sunday', 'Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday']
    if coordinator and not is_admin:
        courses_qs = Course.objects.filter(college=coordinator.college)
        teachers   = Teacher.objects.filter(college=coordinator.college).order_by('name')
        classrooms = Classroom.objects.filter(Q(college=coordinator.college) | Q(college__isnull=True)).order_by('name')
    else:
        courses_qs = Course.objects.all()
        teachers   = Teacher.objects.order_by('name')
        classrooms = Classroom.objects.order_by('name')

    if course_q:
        courses_qs = courses_qs.filter(Q(title__icontains=course_q) | Q(course_code__icontains=course_q))
    # Limit to 200 for performance; user can search to narrow down
    courses = courses_qs.order_by('title')[:200]

    return render(request, 'attendance/add_schedule.html', {
        'courses': courses, 'teachers': teachers, 'classrooms': classrooms,
        'days': DAYS, 'semesters': SEMESTER_CHOICES_4Y,
        'is_admin': is_admin, 'coordinator': coordinator,
    })


@login_required
@require_POST
def delete_schedule(request, schedule_id):
    is_admin    = request.user.is_staff or request.user.is_superuser
    coordinator = Coordinator.objects.filter(auth_user=request.user).first()
    if not (is_admin or coordinator):
        messages.error(request, 'غير مصرح.')
        return redirect('schedule')
    schedule = get_object_or_404(Schedule, pk=schedule_id)
    # Coordinator can only delete schedules in their college
    if coordinator and not is_admin:
        if schedule.course.college != coordinator.college:
            messages.error(request, 'لا يمكنك حذف جداول كليات أخرى.')
            return redirect('schedule')
    schedule.delete()
    messages.success(request, 'تم حذف الحصة.')
    return redirect('schedule')


@login_required
def api_check_conflict(request):
    # Only staff/admin and coordinators may use this endpoint.
    is_admin    = request.user.is_staff or request.user.is_superuser
    coordinator = Coordinator.objects.filter(auth_user=request.user).first()
    if not (is_admin or coordinator):
        return JsonResponse({'error': 'غير مصرح'}, status=403)

    classroom_id = request.GET.get('classroom_id')
    teacher_id   = request.GET.get('teacher_id')
    day          = request.GET.get('day')
    start        = request.GET.get('start')
    end          = request.GET.get('end')
    exclude_id   = request.GET.get('exclude_id')

    base_filter = dict(day_of_week=day, start_time__lt=end, end_time__gt=start)

    # Coordinators may only check resources inside their own college.
    college_scope = coordinator.college if (coordinator and not is_admin) else None

    # Room conflict
    room_conflict = False
    room_msg = ''
    if classroom_id:
        qs = Schedule.objects.filter(classroom_id=classroom_id, **base_filter)
        if college_scope:
            qs = qs.filter(course__college=college_scope)
        if exclude_id:
            qs = qs.exclude(pk=exclude_id)
        if qs.exists():
            room_conflict = True
            room_msg = 'القاعة محجوزة في هذا الوقت'

    # Teacher conflict
    teacher_conflict = False
    teacher_msg = ''
    if teacher_id:
        qs2 = Schedule.objects.filter(teacher_id=teacher_id, **base_filter)
        if college_scope:
            qs2 = qs2.filter(course__college=college_scope)
        if exclude_id:
            qs2 = qs2.exclude(pk=exclude_id)
        if qs2.exists():
            teacher_conflict = True
            teacher_msg = 'الأستاذ مجدول في وقت آخر في نفس اليوم'

    return JsonResponse({
        'conflict': room_conflict or teacher_conflict,
        'room_conflict': room_conflict,
        'room_msg': room_msg,
        'teacher_conflict': teacher_conflict,
        'teacher_msg': teacher_msg,
    })


# ── Courses ──────────────────────────────────────────────────────────────────

@login_required
def courses_list(request):
    coordinator = Coordinator.objects.filter(auth_user=request.user).first()
    is_admin = request.user.is_staff or request.user.is_superuser
    if is_admin:
        courses = Course.objects.select_related('college', 'department').all()
    elif coordinator:
        courses = Course.objects.filter(college=coordinator.college).select_related('college', 'department')
    else:
        courses = Course.objects.none()
    return render(request, 'attendance/courses_list.html', {
        'courses': courses, 'colleges': College.objects.all(),
        'departments': Department.objects.all(),
        'is_admin': is_admin, 'coordinator': coordinator,
    })


@login_required
def add_course(request):
    coordinator = Coordinator.objects.filter(auth_user=request.user).first()
    is_admin = request.user.is_staff or request.user.is_superuser
    if not (is_admin or coordinator):
        messages.error(request, 'غير مصرح.')
        return redirect('courses_list')

    if request.method == 'POST':
        try:
            college_id = request.POST.get('college_id') or (coordinator.college_id if coordinator else None)
            Course.objects.create(
                course_code=request.POST.get('course_code', '').strip(),
                title=request.POST.get('title', '').strip(),
                credits=int(request.POST.get('credits') or 3),
                total_hours=int(request.POST.get('total_hours') or 3),
                college_id=college_id,
                department_id=request.POST.get('department_id') or None,
                year_level=request.POST.get('year_level') or None,
            )
            messages.success(request, 'تمت إضافة المادة.')
            return redirect('courses_list')
        except Exception as e:
            messages.error(request, f'خطأ: {e}')
            # fall through and re-render the form with the error message

    # GET (or POST validation error) → render the add-course form page
    if coordinator and not is_admin:
        departments = Department.objects.filter(college=coordinator.college)
    else:
        departments = Department.objects.all()
    return render(request, 'attendance/add_course.html', {
        'colleges': College.objects.all(),
        'departments': departments,
        'year_choices': range(1, 6),
        'course_types': [('theory', 'نظري'), ('practical', 'عملي'), ('lab', 'معملي'), ('seminar', 'سمنار')],
        'coordinator': coordinator,
        'is_admin': is_admin,
    })


@login_required
def delete_course(request, course_id):
    if not (request.user.is_staff or Coordinator.objects.filter(auth_user=request.user).exists()):
        messages.error(request, 'غير مصرح.')
        return redirect('courses_list')
    Course.objects.filter(pk=course_id).delete()
    messages.success(request, 'تم حذف المادة.')
    return redirect('courses_list')


# ── Classrooms ───────────────────────────────────────────────────────────────

@login_required
def classrooms_list(request):
    from django.core.cache import cache
    from django.db import close_old_connections
    close_old_connections()
    q             = request.GET.get('q', '').strip()
    selected_type = request.GET.get('type', '').strip()
    colleges      = cache.get('colleges_list_qs')
    if colleges is None:
        colleges = list(College.objects.order_by('college_name'))
        cache.set('colleges_list_qs', colleges, 120)
    qs = Classroom.objects.select_related('college').all().order_by('name')
    if q:
        qs = qs.filter(name__icontains=q)
    if selected_type:
        qs = qs.filter(classroom_type=selected_type)
    classrooms = list(qs)
    return render(request, 'attendance/classrooms_list.html', {
        'classrooms': classrooms,
        'types': Classroom.CLASSROOM_TYPES,
        'colleges': colleges,
        'is_admin': request.user.is_staff or request.user.is_superuser,
        'selected_type': selected_type,
        'selected_q': q,
    })


@login_required
@staff_member_required
def add_classroom(request):
    if request.method == 'GET':
        return render(request, 'attendance/add_classroom.html', {
            'types': Classroom.CLASSROOM_TYPES,
            'colleges': College.objects.order_by('college_name'),
        })
    if request.method == 'POST':
        try:
            college_id = request.POST.get('college_id') or None
            Classroom.objects.create(
                name=request.POST.get('name', '').strip(),
                location=request.POST.get('location', '').strip(),
                capacity=int(request.POST.get('capacity', 30)),
                classroom_type=request.POST.get('classroom_type', 'Lecture'),
                college_id=int(college_id) if college_id else None,
            )
            messages.success(request, 'تمت إضافة القاعة.')
        except Exception as e:
            messages.error(request, f'خطأ: {e}')
    return redirect('classrooms_list')


@login_required
@staff_member_required
def delete_classroom(request, classroom_id):
    Classroom.objects.filter(pk=classroom_id).delete()
    messages.success(request, 'تم حذف القاعة.')
    return redirect('classrooms_list')


# ── Register ─────────────────────────────────────────────────────────────────

@login_required
@staff_member_required
def register_student(request):
    is_admin    = request.user.is_superuser or request.user.is_staff
    coordinator = Coordinator.objects.filter(auth_user=request.user).first()
    if not is_admin and not coordinator:
        return _redirect_by_role(request)

    if request.method == 'POST':
        try:
            username     = request.POST.get('username', '').strip()
            password     = request.POST.get('password', 'Shamel@123')
            name         = request.POST.get('name', '').strip()
            student_code = request.POST.get('student_code', '').strip()
            dept_id      = request.POST.get('department_id') or None
            batch        = request.POST.get('batch', '')
            phone        = request.POST.get('phone_number', '').strip()
            email        = request.POST.get('university_email', '').strip()
            blood_type   = request.POST.get('blood_type', '').strip()
            # Nationality: "Sudan" or custom text
            nat_choice   = request.POST.get('nationality_choice', 'Sudan')
            nationality  = 'Sudan' if nat_choice == 'Sudan' else request.POST.get('nationality_text', '').strip()

            # Coordinator can only register in their own college
            if coordinator and not is_admin and dept_id:
                dept = Department.objects.filter(pk=dept_id, college=coordinator.college).first()
                if not dept:
                    messages.error(request, 'لا يمكنك التسجيل في هذا القسم.')
                    return redirect('register_student')

            if not username:
                username = _unique_username(email, student_code, name)
            # Auto-generate unique student_code if blank or already taken
            if not student_code:
                import random as _rnd
                base_code = f'STD{_rnd.randint(100000, 999999)}'
                while Student.objects.filter(student_code=base_code).exists():
                    base_code = f'STD{_rnd.randint(100000, 999999)}'
                student_code = base_code
            elif Student.objects.filter(student_code=student_code).exists():
                messages.error(request, f'رمز الطالب "{student_code}" مستخدم بالفعل — اختر رمزاً آخر.')
                return redirect('register_student')
            user = User.objects.create_user(username=username, password=password, email=email)
            from django.db import connection as _conn
            cursor = _conn.cursor()
            # Insert using raw SQL to pass extra DB-level fields
            cursor.execute("""
                INSERT INTO attendance_student
                  (auth_user_id, student_code, name, department_id, batch,
                   phone_number, university_email, blood_type, nationality,
                   is_registered, is_allowed_entry, is_trained, semester)
                VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,true,true,false,'1')
            """, [user.pk, student_code, name, dept_id, batch,
                  phone or None, email or None,
                  blood_type or None, nationality or None])
            student = Student.objects.get(auth_user=user)
            messages.success(request, f'تم تسجيل الطالب {name}.')
            log_audit(request, 'REGISTER_STUDENT', 'Student', student.pk, name)
        except Exception as e:
            messages.error(request, f'خطأ: {e}')
        return redirect('coordinator_students' if coordinator and not is_admin else 'faculty_management')

    # GET — render form
    if coordinator and not is_admin:
        departments = Department.objects.filter(college=coordinator.college)
        colleges    = College.objects.filter(pk=coordinator.college_id)
    else:
        departments = Department.objects.select_related('college').all()
        colleges    = College.objects.all()
    import datetime as _dt
    current_year = _dt.date.today().year
    return render(request, 'attendance/register_student.html', {
        'departments': departments, 'colleges': colleges,
        'is_admin': is_admin, 'coordinator': coordinator,
        'blood_types': ['A+','A-','B+','B-','AB+','AB-','O+','O-'],
        'batch_years': list(range(current_year, current_year - 10, -1)),
        'semester_choices': SEMESTER_CHOICES_4Y,
    })


@login_required
def register_teacher(request):
    is_admin    = request.user.is_superuser or request.user.is_staff
    coordinator = Coordinator.objects.filter(auth_user=request.user).first()
    if not is_admin and not coordinator:
        return _redirect_by_role(request)

    if request.method == 'POST':
        try:
            username   = request.POST.get('username', '').strip()
            password   = request.POST.get('password', 'Shamel@123')
            name       = request.POST.get('name', '').strip()
            degree     = request.POST.get('academic_degree', 'PhD')
            major      = request.POST.get('major', '').strip()
            college_id = request.POST.get('college_id') or (str(coordinator.college_id) if coordinator else None)
            dept_id    = request.POST.get('department_id') or None
            phone      = request.POST.get('phone_number', '').strip()
            email      = request.POST.get('university_email', '').strip()
            blood_type = request.POST.get('blood_type', '').strip()
            gender     = request.POST.get('gender', 'M')

            # Coordinator scoped to their college
            if coordinator and not is_admin:
                college_id = str(coordinator.college_id)

            if not username:
                username = _unique_username(email, name)
            user = User.objects.create_user(username=username, password=password, is_staff=False, email=email)
            teacher = Teacher.objects.create(
                auth_user=user, name=name, academic_degree=degree,
                major=major, college_id=college_id, department_id=dept_id,
                gender=gender,
                phone_number=phone or None,
                university_email=email or None,
            )
            # blood_type is in DB but not model — set via raw SQL
            if blood_type:
                from django.db import connection as _conn
                _conn.cursor().execute(
                    'UPDATE attendance_teacher SET blood_type=%s WHERE teacher_id=%s',
                    [blood_type, teacher.pk]
                )
            messages.success(request, f'تم تسجيل الأستاذ {name}.')
            log_audit(request, 'REGISTER_TEACHER', 'Teacher', teacher.pk, name)
        except Exception as e:
            messages.error(request, f'خطأ: {e}')
        return redirect('coordinator_faculty' if coordinator and not is_admin else 'faculty_management')

    # GET — render form
    if coordinator and not is_admin:
        departments = Department.objects.filter(college=coordinator.college)
        colleges    = College.objects.filter(pk=coordinator.college_id)
    else:
        departments = Department.objects.select_related('college').all()
        colleges    = College.objects.all()
    return render(request, 'attendance/register_teacher.html', {
        'departments': departments, 'colleges': colleges,
        'is_admin': is_admin, 'coordinator': coordinator,
        'blood_types': ['A+','A-','B+','B-','AB+','AB-','O+','O-'],
        'degrees': ['Prof.','Assoc. Prof.','Asst. Prof.','PhD','MSc','BSc','Lecturer'],
    })


@login_required
@require_POST
def detect_face_api(request):
    """Lightweight endpoint: returns bounding boxes of detected faces for live tracking overlay."""
    if not FACE_ENGINE_AVAILABLE or not CV2_AVAILABLE or not NUMPY_AVAILABLE:
        return JsonResponse({'detected': False, 'faces': []})
    b64 = request.POST.get('frame', '')
    if not b64:
        return JsonResponse({'detected': False, 'faces': []})
    try:
        if ',' in b64:
            b64 = b64.split(',')[1]
        img_bytes = base64.b64decode(b64)
        img_arr = cv2.imdecode(np.frombuffer(img_bytes, dtype=np.uint8), cv2.IMREAD_COLOR)
        if img_arr is None:
            return JsonResponse({'detected': False, 'faces': []})
        h, w = img_arr.shape[:2]
        faces = _fe.detect(img_arr[:, :, ::-1])  # BGR→RGB
        return JsonResponse({
            'detected': len(faces) > 0,
            'faces': faces,
            'width': w,
            'height': h,
        })
    except Exception:
        return JsonResponse({'detected': False, 'faces': []})


@login_required
def enroll_face(request, person_type=None, person_id=None):
    """Enroll a face embedding for the given person (or the logged-in user if no person specified)."""

    # Resolve the target person — URL params take priority over POST fallback
    pt = person_type or request.POST.get('user_type', '')
    pid_raw = person_id or request.POST.get('user_id', '')
    try:
        pid = int(pid_raw) if pid_raw else None
    except (ValueError, TypeError):
        pid = None

    student_obj = None
    teacher_obj = None
    if pt == 'student' and pid:
        student_obj = get_object_or_404(Student, pk=pid)
    elif pt == 'teacher' and pid:
        teacher_obj = get_object_or_404(Teacher, pk=pid)
    else:
        # Self-enrollment: use the logged-in user
        student_obj = Student.objects.filter(auth_user=request.user).first()
        teacher_obj = Teacher.objects.filter(auth_user=request.user).first() if not student_obj else None

    person = student_obj or teacher_obj
    if person is None:
        return HttpResponseBadRequest('لم يتم العثور على الشخص المطلوب')

    user_type_ctx = 'student' if student_obj else 'teacher'

    if request.method == 'POST':
        b64 = request.POST.get('image_data', '')
        if not b64:
            return JsonResponse({'status': 'error', 'message': 'لم تصل صورة'})
        if not FACE_ENGINE_AVAILABLE or not NUMPY_AVAILABLE:
            return JsonResponse({'status': 'error', 'message': 'محرك التعرف على الوجه غير متوفر على الخادم'})
        # angle_index: 0=front (main), 1-4=extra angles
        try:
            angle_index = int(request.POST.get('angle_index', 0))
        except (ValueError, TypeError):
            angle_index = 0
        try:
            if ',' in b64:
                b64 = b64.split(',')[1]
            img_bytes = base64.b64decode(b64)
            import cv2 as _cv2
            img_arr = _cv2.imdecode(
                np.frombuffer(img_bytes, dtype=np.uint8), _cv2.IMREAD_COLOR
            )
            if img_arr is None:
                return JsonResponse({'status': 'error', 'message': 'الصورة غير صالحة'})
            enc = _fe.encode(img_arr[:, :, ::-1])  # BGR→RGB
            if not enc:
                return JsonResponse({'status': 'error', 'message': 'لم يتم اكتشاف وجه — تأكد من الإضاءة وأن الوجه واضح أمام الكاميرا'})
            if student_obj:
                emb_obj, _ = StudentFaceEmbedding.objects.get_or_create(
                    student=student_obj, defaults={'embedding': enc, 'extra_embeddings': []}
                )
                if angle_index == 0:
                    emb_obj.embedding = enc
                else:
                    extras = list(emb_obj.extra_embeddings or [])
                    idx = angle_index - 1
                    if idx < len(extras):
                        extras[idx] = enc
                    else:
                        extras.append(enc)
                    emb_obj.extra_embeddings = extras
                emb_obj.save()
            else:
                emb_obj, _ = TeacherFaceEmbedding.objects.get_or_create(
                    teacher=teacher_obj, defaults={'face_vector': enc, 'extra_embeddings': []}
                )
                if angle_index == 0:
                    emb_obj.face_vector = enc
                else:
                    extras = list(emb_obj.extra_embeddings or [])
                    idx = angle_index - 1
                    if idx < len(extras):
                        extras[idx] = enc
                    else:
                        extras.append(enc)
                    emb_obj.extra_embeddings = extras
                emb_obj.save()
            load_known_faces()
            return JsonResponse({'status': 'success', 'angle_index': angle_index})
        except Exception as e:
            logger.error('enroll_face error: %s', e, exc_info=True)
            return JsonResponse({'status': 'error', 'message': str(e)})

    has_face = (
        StudentFaceEmbedding.objects.filter(student=student_obj).exists()
        if student_obj else
        TeacherFaceEmbedding.objects.filter(teacher=teacher_obj).exists()
    )
    upload_url = reverse('enroll_face_person', args=[user_type_ctx, person.pk])
    return render(request, 'attendance/enroll_face.html', {
        'person': person,
        'user_type': user_type_ctx,
        'user_id': person.pk,
        'has_face': has_face,
        'upload_url': upload_url,
    })


@login_required
def api_departments(request):
    college_id = request.GET.get('college_id', '')
    depts = Department.objects.all()
    if college_id:
        depts = depts.filter(college_id=college_id)
    data = [{'id': d.pk, 'name': d.name} for d in depts]
    return JsonResponse({'departments': data})


# ═══════════════════════════════════════════════════════════════════════════════
# PROFESSOR / TEACHER VIEWS
# ═══════════════════════════════════════════════════════════════════════════════

@login_required
def professor_dashboard(request):
    try:
        teacher = Teacher.objects.get(auth_user=request.user)
    except Teacher.DoesNotExist:
        return _redirect_by_role(request)
    if teacher is None:
        return _redirect_by_role(request)
    schedules      = Schedule.objects.filter(teacher=teacher).select_related('course', 'classroom')
    active_session = LectureSession.objects.filter(
        schedule__teacher=teacher, is_active=True
    ).select_related('schedule__course', 'schedule__classroom').first()
    today          = timezone.now().date()
    now            = timezone.now()

    # Attendance logs for today's active session
    present_students = []
    absent_students  = []
    present_count    = 0
    total_count      = 0

    if active_session:
        session_logs = AIAttendanceLog.objects.filter(
            session=active_session
        ).select_related('student').order_by('-timestamp')
        present_students = session_logs.filter(status='Present')
        present_count    = present_students.count()
        # Absent = enrolled but not in present list
        enrolled_ids     = Enrollment.objects.filter(
            course=active_session.schedule.course
        ).values_list('student_id', flat=True)
        present_ids      = present_students.values_list('student_id', flat=True)
        absent_students  = Student.objects.filter(
            id__in=enrolled_ids
        ).exclude(id__in=present_ids)
        total_count      = len(enrolled_ids)

    # Recent attendance logs for today (for teacher's courses)
    recent_logs = AIAttendanceLog.objects.filter(
        schedule__teacher=teacher, timestamp__date=today
    ).select_related('student').order_by('-timestamp')[:10]

    # Next upcoming schedule today
    next_session = Schedule.objects.filter(
        teacher=teacher,
        day_of_week=now.strftime('%A'),
        start_time__gt=now.time(),
    ).select_related('course', 'classroom').order_by('start_time').first()

    # Classrooms for room-availability sidebar
    classrooms = Classroom.objects.order_by('name')[:12]

    return render(request, 'attendance/professor_dashboard.html', {
        'teacher':          teacher,
        'schedules':        schedules,
        'active_session':   active_session,
        'current_session':  active_session,   # alias used by template
        'recent_logs':      recent_logs,
        'present_students': present_students,
        'absent_students':  absent_students,
        'present_count':    present_count,
        'total_count':      total_count,
        'next_session':     next_session,
        'classrooms':       classrooms,
    })


@login_required
def open_session(request, schedule_id):
    teacher = get_object_or_404(Teacher, auth_user=request.user)
    schedule = get_object_or_404(Schedule, pk=schedule_id, teacher=teacher)
    # Close any existing active session for this schedule
    LectureSession.objects.filter(schedule=schedule, is_active=True).update(
        is_active=False, actual_end_time=timezone.now()
    )
    session = LectureSession.objects.create(
        schedule=schedule, is_active=True,
        actual_start_time=timezone.now(), opened_by=request.user,
    )
    log_audit(request, 'OPEN_SESSION', 'LectureSession', session.pk,
              f'Teacher {teacher.name} opened session for {schedule}')
    messages.success(request, 'تم فتح الجلسة.')
    return redirect('lecture_scan', session_id=session.pk)


@login_required
def teacher_timeline(request):
    try:
        teacher = Teacher.objects.get(auth_user=request.user)
    except Teacher.DoesNotExist:
        return _redirect_by_role(request)
    sessions = LectureSession.objects.filter(
        schedule__teacher=teacher
    ).select_related('schedule__course').order_by('-actual_start_time')[:50]
    return render(request, 'attendance/teacher_timeline.html', {
        'teacher': teacher, 'sessions': sessions,
    })


# ── Lecture Scan (attendance scan tied to a LectureSession) ─────────────────

@login_required
def lecture_scan(request, session_id):
    teacher = get_object_or_404(Teacher, auth_user=request.user)
    session = get_object_or_404(LectureSession, pk=session_id,
                                 is_active=True, schedule__teacher=teacher)
    enrolled = (Enrollment.objects.filter(course=session.schedule.course)
                .select_related('student').order_by('student__name'))
    present_ids = set(AIAttendanceLog.objects.filter(
        session=session, status='Present').values_list('student_id', flat=True))
    return render(request, 'attendance/lecture_scan.html', {
        'session': session,
        'enrolled': enrolled,
        'present_ids': present_ids,
        'face_engine_available': FACE_ENGINE_AVAILABLE and CV2_AVAILABLE and NUMPY_AVAILABLE,
    })


@login_required
@csrf_exempt
def lecture_scan_api(request):
    """POST: frame (base64) + session_id → match enrolled students → log AIAttendanceLog."""
    if request.method != 'POST':
        return JsonResponse({'ok': False}, status=405)
    teacher = get_object_or_404(Teacher, auth_user=request.user)
    session_id = request.POST.get('session_id')
    if not session_id:
        return JsonResponse({'ok': False, 'error': 'no_session_id'})
    session = get_object_or_404(LectureSession, pk=session_id,
                                 is_active=True, schedule__teacher=teacher)

    if not (FACE_ENGINE_AVAILABLE and CV2_AVAILABLE and NUMPY_AVAILABLE):
        return JsonResponse({'ok': False, 'error': 'engine_unavailable'})

    b64 = request.POST.get('frame', '')
    if not b64:
        return JsonResponse({'ok': False, 'error': 'no_frame'})

    try:
        if ',' in b64:
            b64 = b64.split(',')[1]
        img_bytes = base64.b64decode(b64)
        img_arr = cv2.imdecode(np.frombuffer(img_bytes, dtype=np.uint8), cv2.IMREAD_COLOR)
        if img_arr is None:
            return JsonResponse({'ok': False, 'error': 'bad_image'})

        import time as _time
        all_faces = _fe.encode_all(img_arr[:, :, ::-1])

        if not all_faces:
            return JsonResponse({'ok': True, 'detected': False, 'results': [],
                                 'width': img_arr.shape[1], 'height': img_arr.shape[0]})

        enrolled_ids = set(Enrollment.objects.filter(
            course=session.schedule.course).values_list('student_id', flat=True))
        now_ts = _time.time()
        results = []

        for face_data in all_faces:
            bbox = face_data['bbox']
            matched_name, matched_type, matched_pk = match_face_from_db(face_data['embedding'])
            if not matched_name:
                results.append({'bbox': bbox, 'matched': False})
                continue

            # Teacher face → acknowledge but don't log as student
            if matched_type == 'teacher':
                results.append({'bbox': bbox, 'matched': True,
                                 'name': matched_name, 'is_teacher': True, 'logged': False})
                continue

            if matched_pk not in enrolled_ids:
                results.append({'bbox': bbox, 'matched': False, 'not_enrolled': True})
                continue

            student = Student.objects.filter(pk=matched_pk).first()
            if not student:
                results.append({'bbox': bbox, 'matched': False})
                continue

            cooldown_key = f'lscan_{session_id}_{matched_pk}'
            with _gate_cooldown_lock:
                last = _gate_cooldown.get(cooldown_key, 0)
                in_cooldown = (now_ts - last) < 30
                if not in_cooldown:
                    _gate_cooldown[cooldown_key] = now_ts

            already_logged = AIAttendanceLog.objects.filter(
                session=session, student=student).exists()

            if not in_cooldown and not already_logged:
                AIAttendanceLog.objects.create(
                    student=student,
                    schedule=session.schedule,
                    session=session,
                    status='Present',
                    method='face_recognition',
                )

            results.append({
                'name': matched_name,
                'type': 'student',
                'matched': True,
                'already_logged': already_logged,
                'bbox': bbox,
            })

        return JsonResponse({
            'ok': True, 'detected': True, 'results': results,
            'width': img_arr.shape[1], 'height': img_arr.shape[0],
        })
    except Exception as exc:
        return JsonResponse({'ok': False, 'error': str(exc)[:200]})


@login_required
def lecture_session_status(request, session_id):
    """GET → JSON of present/absent students for live polling."""
    teacher = get_object_or_404(Teacher, auth_user=request.user)
    session = get_object_or_404(LectureSession, pk=session_id, schedule__teacher=teacher)
    enrolled = list(Enrollment.objects.filter(
        course=session.schedule.course).select_related('student'))
    logs = {l.student_id: l for l in AIAttendanceLog.objects.filter(
        session=session, status='Present').select_related('student')}
    present = [{'id': e.student.id, 'name': e.student.name,
                'time': logs[e.student_id].timestamp.strftime('%H:%M')}
               for e in enrolled if e.student_id in logs]
    absent = [{'id': e.student.id, 'name': e.student.name}
              for e in enrolled if e.student_id not in logs]
    return JsonResponse({'present': present, 'absent': absent,
                         'total': len(enrolled)})


@login_required
def lecture_manual_attendance(request, session_id):
    """POST present_ids[] → bulk upsert AIAttendanceLog (offline fallback)."""
    teacher = get_object_or_404(Teacher, auth_user=request.user)
    session = get_object_or_404(LectureSession, pk=session_id, schedule__teacher=teacher)
    enrolled = (Enrollment.objects.filter(course=session.schedule.course)
                .select_related('student').order_by('student__name'))
    if request.method == 'POST':
        present_ids = set(request.POST.getlist('present_ids'))
        for e in enrolled:
            AIAttendanceLog.objects.update_or_create(
                student=e.student, session=session,
                defaults={
                    'schedule': session.schedule,
                    'status': 'Present' if str(e.student_id) in present_ids else 'Absent',
                    'method': 'manual',
                },
            )
        messages.success(request, 'تم تسجيل الحضور يدوياً.')
        if session.is_active:
            return redirect('lecture_scan', session_id=session_id)
        return redirect('professor_dashboard')
    present_ids = set(AIAttendanceLog.objects.filter(
        session=session, status='Present').values_list('student_id', flat=True))
    return render(request, 'attendance/lecture_manual_attendance.html', {
        'session': session, 'enrolled': enrolled, 'present_ids': present_ids,
    })


@login_required
def export_my_courses_csv(request):
    teacher = get_object_or_404(Teacher, auth_user=request.user)
    schedules = Schedule.objects.filter(teacher=teacher).select_related('course', 'classroom')
    response = HttpResponse(content_type='text/csv; charset=utf-8-sig')
    from datetime import date
    tname = teacher.name.replace(' ', '_')[:30]
    response['Content-Disposition'] = f'attachment; filename="SHAMEL_{tname}_MyCourses_{date.today()}.csv"'
    writer = csv.writer(response)
    writer.writerow(['Course Code', 'Title', 'Day', 'Start', 'End', 'Classroom', 'Batch', 'Semester'])
    for s in schedules:
        writer.writerow([
            s.course.course_code if s.course else '',
            s.course.title if s.course else '',
            s.day_of_week, s.start_time, s.end_time,
            s.classroom.name if s.classroom else '',
            s.batch, s.semester,
        ])
    return response


@login_required
def teacher_attendance_report(request):
    from datetime import date as _date
    teacher_q  = request.GET.get('teacher_q', '').strip()
    date_from  = request.GET.get('date_from', '')
    date_to    = request.GET.get('date_to', '')

    teachers_qs = Teacher.objects.select_related('department', 'college').order_by('name')
    if teacher_q:
        teachers_qs = teachers_qs.filter(name__icontains=teacher_q)

    sessions_qs = LectureSession.objects.filter(is_active=False).select_related(
        'schedule__teacher', 'schedule__course', 'schedule__classroom'
    )
    if date_from:
        try:
            from datetime import datetime as _dt
            sessions_qs = sessions_qs.filter(actual_start_time__date__gte=_dt.strptime(date_from, '%Y-%m-%d').date())
        except ValueError:
            pass
    if date_to:
        try:
            from datetime import datetime as _dt
            sessions_qs = sessions_qs.filter(actual_start_time__date__lte=_dt.strptime(date_to, '%Y-%m-%d').date())
        except ValueError:
            pass
    if teacher_q:
        sessions_qs = sessions_qs.filter(schedule__teacher__name__icontains=teacher_q)

    # ── N+1 fix: group sessions by teacher_id in Python once (O(S)) ──
    from collections import defaultdict
    session_list = list(sessions_qs)   # single DB hit
    by_teacher   = defaultdict(list)
    for s in session_list:
        if s.schedule_id:
            by_teacher[s.schedule.teacher_id].append(s)

    data = []
    for t in teachers_qs:
        t_sessions    = by_teacher.get(t.teacher_id, [])
        total_minutes = sum((s.duration_minutes or 0) for s in t_sessions)
        data.append({
            'teacher':       t,
            'sessions':      len(t_sessions),
            'total_minutes': total_minutes,
            'session_list':  t_sessions[:5],
        })

    return render(request, 'attendance/reports/teacher_report.html', {
        'data':           data,
        'teacher_q':      teacher_q,
        'date_from':      date_from,
        'date_to':        date_to,
        'total_sessions': sum(d['sessions'] for d in data),
    })


@login_required
@staff_member_required
def export_teacher_report_csv(request):
    teachers = Teacher.objects.select_related('department', 'college').order_by('name')
    response = HttpResponse(content_type='text/csv; charset=utf-8-sig')
    from datetime import date
    response['Content-Disposition'] = f'attachment; filename="SHAMEL_TeacherAttendance_Report_{date.today()}.csv"'
    writer = csv.writer(response)
    writer.writerow(['Name', 'Degree', 'Major', 'College', 'Department', 'Sessions Held'])
    for t in teachers:
        sessions = LectureSession.objects.filter(schedule__teacher=t, is_active=False).count()
        writer.writerow([
            t.name, t.academic_degree, t.major,
            t.college.college_name if t.college else '',
            t.department.name if t.department else '',
            sessions,
        ])
    return response


# ═══════════════════════════════════════════════════════════════════════════════
# STUDENT VIEWS
# ═══════════════════════════════════════════════════════════════════════════════

@login_required
def student_dashboard(request):
    student = get_object_or_404(Student, auth_user=request.user)
    today = timezone.now().date()
    logs = AIAttendanceLog.objects.filter(student=student).order_by('-timestamp')
    total = logs.count()
    present = logs.filter(status='Present').count()
    pct = round(present / total * 100, 1) if total else 0
    recent = logs[:5]
    enrollments = Enrollment.objects.filter(student=student).select_related('course')
    notifications = Notification.objects.filter(user=request.user, is_read=False)[:5]
    return render(request, 'attendance/student_dashboard.html', {
        'student': student, 'total': total, 'present': present, 'pct': pct,
        'recent_logs': recent, 'enrollments': enrollments,
        'notifications': notifications,
    })


@login_required
def student_profile(request):
    student = get_object_or_404(Student, auth_user=request.user)
    # gender, nationality, blood_type are DB-only fields (not in model)
    gender_val = '—'
    nationality_val = ''
    blood_type_val = ''
    try:
        with connection.cursor() as cur:
            cur.execute(
                'SELECT gender, nationality, blood_type FROM attendance_student WHERE id=%s',
                [student.pk]
            )
            row = cur.fetchone()
            if row:
                g, nat, bt = row
                GENDER_MAP = {'M': 'ذكر', 'F': 'أنثى'}
                gender_val = GENDER_MAP.get(g or '', g or '—') or '—'
                nationality_val = nat or ''
                blood_type_val = bt or ''
    except Exception:
        pass
    college_name = '—'
    try:
        if student.department and student.department.college:
            college_name = student.department.college.college_name
    except Exception:
        pass
    student_fields = [
        ('الاسم', student.name),
        ('رقم الطالب', student.student_code),
        ('القسم', student.department.name if student.department else '—'),
        ('الكلية', college_name),
        ('الجنس', gender_val),
        ('الدفعة', student.batch or '—'),
        ('الجنسية', nationality_val or '—'),
        ('فصيلة الدم', blood_type_val or '—'),
    ]
    return render(request, 'attendance/student_profile.html', {
        'student': student,
        'student_fields': student_fields,
    })


@login_required
def student_courses(request):
    student = get_object_or_404(Student, auth_user=request.user)

    # My enrolled courses
    enrollments = Enrollment.objects.filter(student=student).select_related('course', 'classroom', 'course__department')
    enrolled_ids = set(enrollments.values_list('course_id', flat=True))

    # Student's own department + college (never show other departments)
    student_dept   = student.department
    student_college = student_dept.college if student_dept else None

    # Filters
    semester_filter = request.GET.get('semester', '')
    search_q        = request.GET.get('q', '').strip()

    # Base: only student's department courses (or college if no dept)
    if student_dept:
        all_courses_qs = Course.objects.filter(department=student_dept).select_related('department', 'college')
    elif student_college:
        all_courses_qs = Course.objects.filter(college=student_college).select_related('department', 'college')
    else:
        all_courses_qs = Course.objects.none()

    if search_q:
        all_courses_qs = all_courses_qs.filter(Q(title__icontains=search_q) | Q(course_code__icontains=search_q))

    if semester_filter:
        # filter by year_level matching semester (semester / 2 rounds up = year)
        try:
            sem_int = int(semester_filter)
            year_level = (sem_int + 1) // 2
            all_courses_qs = all_courses_qs.filter(year_level=year_level)
        except ValueError:
            pass

    # Attach course_type from DB
    course_type_map = {}
    try:
        with connection.cursor() as cur:
            cur.execute("SELECT id, course_type FROM attendance_course")
            course_type_map = {r[0]: r[1] for r in cur.fetchall()}
    except Exception:
        pass

    all_courses = list(all_courses_qs.order_by('year_level', 'title'))
    for c in all_courses:
        c.course_type_val = course_type_map.get(c.pk, '')

    # "Current Courses" — schedules the coordinator created for student's dept this semester
    current_schedules = []
    try:
        current_schedules = Schedule.objects.filter(
            course__department=student_dept
        ).select_related('course', 'teacher', 'classroom').order_by('day_of_week', 'start_time') if student_dept else []
    except Exception:
        pass

    return render(request, 'attendance/student_courses.html', {
        'student':          student,
        'enrollments':      enrollments,
        'enrolled_ids':     enrolled_ids,
        'all_courses':      all_courses,
        'semester_filter':  semester_filter,
        'semester_choices': SEMESTER_CHOICES_4Y,
        'search_q':         search_q,
        'current_schedules': current_schedules,
        'student_dept':     student_dept,
    })


@login_required
def student_support(request):
    try:
        student = Student.objects.get(auth_user=request.user)
    except Student.DoesNotExist:
        return _redirect_by_role(request)
    tickets = SupportTicket.objects.filter(user=request.user).order_by('-created_at')
    return render(request, 'attendance/student_support.html', {
        'student': student, 'tickets': tickets,
    })


# ═══════════════════════════════════════════════════════════════════════════════
# COORDINATOR VIEWS
# ═══════════════════════════════════════════════════════════════════════════════

@login_required
def coordinator_dashboard(request):
    coordinator    = get_object_or_404(Coordinator, auth_user=request.user)
    students_count = Student.objects.filter(department__college=coordinator.college).count()
    teachers_count = Teacher.objects.filter(college=coordinator.college).count()
    courses_count  = Course.objects.filter(college=coordinator.college).count()
    today          = timezone.now().date()
    attendance_today = AIAttendanceLog.objects.filter(
        schedule__course__college=coordinator.college, timestamp__date=today
    ).count()
    # Coordinator sees college-scoped tickets, not global system tickets
    open_tickets = SupportTicket.objects.filter(
        status='open',
        user__student__department__college=coordinator.college
    ).count() + SupportTicket.objects.filter(
        status='open',
        user__teacher__college=coordinator.college
    ).count()

    # Academic-specific KPIs (not shown to admin)
    pending_excuses  = MedicalExcuse.objects.filter(
        student__department__college=coordinator.college,
        status='pending'
    ).count()

    # College-wide attendance percentage (last 30 days)
    from datetime import timedelta
    month_ago = today - timedelta(days=30)
    logs_qs   = AIAttendanceLog.objects.filter(
        schedule__course__college=coordinator.college,
        timestamp__date__gte=month_ago,
    )
    total_logs   = logs_qs.count() or 1
    present_logs = logs_qs.filter(status='Present').count()
    college_attendance_pct = round(present_logs / total_logs * 100, 1)

    # Students with attendance below 75% (at-risk)
    from collections import defaultdict
    student_buckets = defaultdict(lambda: {'present': 0, 'total': 0, 'student': None})
    for log in logs_qs.select_related('student'):
        sid = log.student_id
        student_buckets[sid]['student'] = log.student
        student_buckets[sid]['total']  += 1
        if log.status == 'Present':
            student_buckets[sid]['present'] += 1
    low_attendance_students = []
    for data in student_buckets.values():
        rate = round(data['present'] / data['total'] * 100, 1) if data['total'] else 0
        if rate < 75:
            low_attendance_students.append({'student': data['student'], 'rate': rate})
    low_attendance_students.sort(key=lambda x: x['rate'])

    return render(request, 'attendance/coordinator_dashboard.html', {
        'coordinator':            coordinator,
        'college_name':           coordinator.college.college_name if coordinator.college else '',
        # Academic stats (college-scoped only)
        'students_count':         students_count,
        'teachers_count':         teachers_count,
        'courses_count':          courses_count,
        'attendance_today':       attendance_today,
        'open_tickets':           open_tickets,
        'total_students':         students_count,
        'total_teachers':         teachers_count,
        'college_attendance_pct': college_attendance_pct,
        # Academic KPIs (coordinator-specific, not shown to admin)
        'pending_excuses':        pending_excuses,
        'low_attendance_students': low_attendance_students,
    })


@login_required
def coordinator_students(request):
    try:
        coordinator = Coordinator.objects.get(auth_user=request.user)
    except Coordinator.DoesNotExist:
        return _redirect_by_role(request)
    students = Student.objects.filter(
        department__college=coordinator.college
    ).select_related('department', 'department__college', 'auth_user').order_by('name')
    q = request.GET.get('q', '')
    dept_f = request.GET.get('dept', '')
    if q:
        students = students.filter(name__icontains=q)
    if dept_f:
        students = students.filter(department_id=dept_f)
    departments = Department.objects.filter(college=coordinator.college)
    return render(request, 'attendance/coordinator_students.html', {
        'coordinator': coordinator, 'students': students,
        'departments': departments, 'filters': {'q': q, 'dept': dept_f},
    })


@login_required
def coordinator_faculty(request):
    coordinator = get_object_or_404(Coordinator, auth_user=request.user)
    teachers = Teacher.objects.filter(college=coordinator.college).select_related('department').order_by('name')
    q = request.GET.get('q', '')
    if q:
        teachers = teachers.filter(name__icontains=q)
    return render(request, 'attendance/coordinator_faculty.html', {
        'coordinator': coordinator, 'teachers': teachers, 'q': q,
    })


@login_required
def coordinator_course_assignment(request):
    coordinator = get_object_or_404(Coordinator, auth_user=request.user)
    schedules = Schedule.objects.filter(
        course__college=coordinator.college
    ).select_related('course', 'teacher', 'classroom').order_by('day_of_week', 'start_time')

    if request.method == 'POST':
        schedule_id = request.POST.get('schedule_id')
        teacher_id = request.POST.get('teacher_id')
        if schedule_id and teacher_id:
            sch = get_object_or_404(Schedule, pk=schedule_id)
            sch.teacher = get_object_or_404(Teacher, pk=teacher_id)
            sch.save(update_fields=['teacher'])
            messages.success(request, 'تم تحديث تعيين الأستاذ.')
        return redirect('coordinator_course_assignment')

    teachers = Teacher.objects.filter(college=coordinator.college)
    return render(request, 'attendance/coordinator_assignments.html', {
        'coordinator': coordinator, 'schedules': schedules, 'teachers': teachers,
    })


@login_required
def coordinator_register_user(request):
    # Allow admin/staff to reach this page (they see all colleges)
    coordinator = Coordinator.objects.filter(auth_user=request.user).first()
    if coordinator is None and not (request.user.is_staff or request.user.is_superuser):
        return _redirect_by_role(request)
    if coordinator is None:
        # Admin acting as coordinator — create a stub for template compat
        coordinator = type('FakeCoord', (), {
            'college': None,
            'id': None,
        })()
    if request.method == 'POST':
        user_type = request.POST.get('user_type', 'student')
        username = request.POST.get('username', '').strip()
        password = request.POST.get('password', 'Shamel@123')
        name = request.POST.get('name', '').strip()
        email = request.POST.get('email', '').strip()
        student_code = request.POST.get('student_code', '').strip()
        # The form collects name/email/password (no username field) → derive a
        # unique username from email local-part / student code / name.
        if not username:
            base = ''
            if email and '@' in email:
                base = email.split('@')[0]
            base = base or student_code or (name.replace(' ', '_').lower() if name else '')
            base = re.sub(r'[^a-zA-Z0-9_.]', '', base) or 'user'
            username = base
            n = 1
            while User.objects.filter(username=username).exists():
                n += 1
                username = f'{base}{n}'
        try:
            if User.objects.filter(username=username).exists():
                messages.error(request, 'اسم المستخدم مستخدم بالفعل.')
                return redirect('coordinator_register_user')
            user = User.objects.create_user(username=username, password=password, email=email)
            if user_type == 'student':
                dept_id = request.POST.get('department_id') or None
                # Auto-generate unique student_code if blank or duplicate
                if not student_code:
                    import random as _rnd2
                    student_code = f'STD{_rnd2.randint(100000, 999999)}'
                    while Student.objects.filter(student_code=student_code).exists():
                        student_code = f'STD{_rnd2.randint(100000, 999999)}'
                elif Student.objects.filter(student_code=student_code).exists():
                    user.delete()
                    messages.error(request, f'رمز الطالب "{student_code}" مستخدم بالفعل.')
                    return redirect('coordinator_register_user')
                Student.objects.create(
                    auth_user=user, name=name, student_code=student_code,
                    department_id=dept_id, university_email=email,
                )
            elif user_type == 'teacher':
                Teacher.objects.create(
                    auth_user=user, name=name,
                    college=coordinator.college,
                    academic_degree=request.POST.get('academic_degree', 'PhD'),
                )
            messages.success(request, f'تم تسجيل {name}.')
            log_audit(request, f'REGISTER_{user_type.upper()}', user_type.capitalize(), user.pk, name)
        except Exception as e:
            messages.error(request, f'خطأ: {e}')
        return redirect('coordinator_register_user')

    departments = Department.objects.filter(college=coordinator.college) if coordinator.college else Department.objects.all()
    import datetime
    current_year = datetime.date.today().year
    batch_years = list(range(current_year, current_year - 10, -1))
    semester_choices = SEMESTER_CHOICES_4Y
    return render(request, 'attendance/coordinator_register.html', {
        'coordinator': coordinator, 'departments': departments,
        'batch_years': batch_years, 'semester_choices': semester_choices,
    })


# ── PDF exports ──────────────────────────────────────────────────────────────

def _pdf_logo_url(request):
    """Absolute URL to the dark logo for PDF headers (WeasyPrint needs absolute URLs)."""
    from django.templatetags.static import static as _static
    return request.build_absolute_uri(_static('images/logo_dark.svg'))


def _pdf_response(html_string, filename, base_url=None):
    """Render HTML → PDF. Tries weasyprint (best CSS, Linux/VPS), then
    xhtml2pdf (pure-Python, works on Windows), then a printable-HTML fallback."""
    # 1) weasyprint — full CSS, needs GTK/cairo (usually only on Linux VPS)
    try:
        from weasyprint import HTML as WeasyprintHTML
        pdf_bytes = WeasyprintHTML(string=html_string, base_url=base_url).write_pdf()
        resp = HttpResponse(pdf_bytes, content_type='application/pdf')
        resp['Content-Disposition'] = f'attachment; filename="{filename}"'
        return resp
    except Exception:
        pass
    # 2) xhtml2pdf — pure-Python (reportlab), cross-platform.
    #    Its CSS parser chokes on @page running-content rules and some modern
    #    CSS, so sanitize first: drop @page{} blocks and unsupported props.
    try:
        from xhtml2pdf import pisa
        import io as _io, re as _re
        safe = html_string
        # Remove @page {...} blocks (xhtml2pdf mishandles content:/running())
        safe = _re.sub(r'@page[^{]*\{[^}]*\}', '', safe)
        # Remove content: "..." declarations anywhere (running headers/footers)
        safe = _re.sub(r'content\s*:\s*"[^"]*"\s*;?', '', safe)
        buf = _io.BytesIO()
        result = pisa.CreatePDF(src=safe, dest=buf, encoding='utf-8')
        if not result.err and buf.getvalue()[:4] == b'%PDF':
            resp = HttpResponse(buf.getvalue(), content_type='application/pdf')
            resp['Content-Disposition'] = f'attachment; filename="{filename}"'
            return resp
    except Exception:
        pass
    # 3) Fallback: printable HTML with auto-print trigger (browser → Save as PDF)
    printable = html_string.replace(
        '</body>',
        '<script>window.onload=function(){setTimeout(function(){window.print();},300);};</script></body>'
    )
    resp = HttpResponse(printable, content_type='text/html; charset=utf-8')
    return resp


@login_required
def export_student_report_pdf(request):
    from datetime import date as _date
    logs, filters, meta = _build_attendance_queryset(request)
    summary = _summarise_logs(logs)

    # Per-course attendance chart data
    course_stats = {}
    for log in logs:
        key = getattr(getattr(log, 'schedule', None), 'course', None)
        if key:
            title = key.title
            if title not in course_stats:
                course_stats[title] = {'present': 0, 'absent': 0, 'total': 0}
            course_stats[title]['total'] += 1
            if log.status == 'Present':
                course_stats[title]['present'] += 1
            elif log.status in ('Absent', 'Late'):
                course_stats[title]['absent'] += 1

    # Build percentage data for SVG bar chart (max 10 courses)
    chart_data = []
    for title, stats in list(course_stats.items())[:10]:
        pct = round(stats['present'] / stats['total'] * 100) if stats['total'] else 0
        chart_data.append({'title': title[:25], 'pct': pct, 'total': stats['total']})

    # Overall stats
    total     = len(logs)
    present   = sum(1 for l in logs if l.status == 'Present')
    absent    = sum(1 for l in logs if l.status == 'Absent')
    late      = sum(1 for l in logs if l.status == 'Late')
    pct_overall = round(present / total * 100) if total else 0

    html = render(request, 'attendance/reports/student_report_pdf.html', {
        'summary': summary, 'filters': filters,
        'logs': logs, 'chart_data': chart_data,
        'total': total, 'present': present, 'absent': absent, 'late': late,
        'pct_overall': pct_overall,
        'pct_remaining': 100 - pct_overall,
        'generated_at': timezone.now(),
        'logo_url': _pdf_logo_url(request),
        **meta,
    }).content.decode('utf-8')
    # Notify admins about report generation
    try:
        from .email_utils import notify_admin_new_report
        for admin_user in User.objects.filter(is_staff=True).exclude(email=''):
            notify_admin_new_report(admin_user.email, 'تقرير حضور الطلاب', request.user.get_full_name() or request.user.username)
    except Exception:
        pass
    return _pdf_response(html, f'SHAMEL_StudentReport_{_date.today()}.pdf', base_url=request.build_absolute_uri('/'))


@login_required
def export_teacher_report_pdf(request):
    from datetime import date as _date
    teachers = Teacher.objects.select_related('department', 'college').order_by('name')
    data = []
    max_sessions = 1
    total_present_all = 0
    total_att_all = 0
    for t in teachers:
        sessions = LectureSession.objects.filter(schedule__teacher=t, is_active=False).count()
        present_count = AIAttendanceLog.objects.filter(schedule__teacher=t, status='Present').count()
        absent_count  = AIAttendanceLog.objects.filter(schedule__teacher=t, status='Absent').count()
        att_total = present_count + absent_count
        att_pct   = round(present_count / att_total * 100) if att_total else 0
        sessions_pct = 0  # filled after max_sessions known
        total_present_all += present_count
        total_att_all     += att_total
        data.append({
            'teacher': t,
            'sessions': sessions,
            'present': present_count,
            'absent': absent_count,
            'att_total': att_total,
            'att_pct': att_pct,
            'sessions_pct': 0,
        })
        if sessions > max_sessions:
            max_sessions = sessions
    # Fill sessions_pct now that max_sessions is known
    for item in data:
        item['sessions_pct'] = round(item['sessions'] / max_sessions * 100) if max_sessions else 0
    overall_pct = round(total_present_all / total_att_all * 100) if total_att_all else 0
    html = render(request, 'attendance/reports/teacher_report_pdf.html', {
        'data': data,
        'max_sessions': max_sessions,
        'total_student_records': total_att_all,
        'overall_pct': overall_pct,
        'generated_at': timezone.now(),
        'logo_url': _pdf_logo_url(request),
    }).content.decode('utf-8')
    # Notify admins about report generation
    try:
        from .email_utils import notify_admin_new_report
        for admin_user in User.objects.filter(is_staff=True).exclude(email=''):
            notify_admin_new_report(admin_user.email, 'تقرير حضور الأساتذة', request.user.get_full_name() or request.user.username)
    except Exception:
        pass
    return _pdf_response(html, f'SHAMEL_TeacherReport_{_date.today()}.pdf', base_url=request.build_absolute_uri('/'))


@login_required
def export_analytics_pdf(request):
    from datetime import date as _date
    today = timezone.now().date()
    week_ago = today - timedelta(days=6)

    # Overall stats
    total_students   = Student.objects.count()
    total_teachers   = Teacher.objects.count()
    total_courses    = Course.objects.count()
    total_sessions   = LectureSession.objects.filter(is_active=False).count()
    total_attendance = AIAttendanceLog.objects.count()

    present_count = AIAttendanceLog.objects.filter(status='Present').count()
    absent_count  = AIAttendanceLog.objects.filter(status='Absent').count()
    late_count    = AIAttendanceLog.objects.filter(status='Late').count()
    overall_present_pct = round(present_count / total_attendance * 100) if total_attendance else 0

    # Weekly chart (last 7 days)
    DAY_LABELS = {'Saturday':'سبت','Sunday':'أحد','Monday':'اثن','Tuesday':'ثلث','Wednesday':'أرب','Thursday':'خمس','Friday':'جمع'}
    weekly_chart = []
    for i in range(7):
        d = week_ago + timedelta(days=i)
        day_total   = AIAttendanceLog.objects.filter(timestamp__date=d).count()
        day_present = AIAttendanceLog.objects.filter(timestamp__date=d, status='Present').count()
        pct = round(day_present / day_total * 100) if day_total else 0
        weekly_chart.append({
            'label': DAY_LABELS.get(d.strftime('%A'), d.strftime('%a')),
            'count': day_total,
            'pct': pct,
        })

    # Department chart
    dept_chart = []
    for dept in Department.objects.all():
        students = Student.objects.filter(department=dept)
        tot = AIAttendanceLog.objects.filter(student__in=students).count()
        pre = AIAttendanceLog.objects.filter(student__in=students, status='Present').count()
        pct = round(pre / tot * 100) if tot else 0
        dept_chart.append({'name': dept.name, 'pct': pct, 'present': pre, 'total': tot})
    dept_chart = sorted(dept_chart, key=lambda x: -x['pct'])[:8]

    # College chart
    college_chart = []
    for col in College.objects.all():
        students_count = Student.objects.filter(department__college=col).count()
        tot = AIAttendanceLog.objects.filter(student__department__college=col).count()
        pre = AIAttendanceLog.objects.filter(student__department__college=col, status='Present').count()
        pct = round(pre / tot * 100) if tot else 0
        college_chart.append({'name': col.college_name, 'pct': pct, 'students': students_count})
    college_chart = sorted(college_chart, key=lambda x: -x['pct'])[:6]

    # Top courses by session count
    from django.db.models import Count as DBCount
    course_sessions = (LectureSession.objects.filter(is_active=False)
                       .values('schedule__course__title', 'schedule__course__department__name')
                       .annotate(cnt=DBCount('id')).order_by('-cnt')[:8])
    max_cs = max((x['cnt'] for x in course_sessions), default=1)
    top_courses = [{
        'title': x['schedule__course__title'] or '—',
        'dept':  x['schedule__course__department__name'] or '—',
        'sessions': x['cnt'],
        'pct': round(x['cnt'] / max_cs * 100),
    } for x in course_sessions]

    # Top students by attendance
    top_students_qs = (AIAttendanceLog.objects
                       .values('student__id', 'student__name', 'student__student_code',
                               'student__department__name')
                       .annotate(total=DBCount('id'),
                                 present=DBCount('id', filter=Q(status='Present')))
                       .order_by('-present')[:10])
    top_students = [{
        'name': x['student__name'],
        'code': x['student__student_code'],
        'dept': x['student__department__name'] or '—',
        'present': x['present'],
        'pct': round(x['present'] / x['total'] * 100) if x['total'] else 0,
    } for x in top_students_qs]

    html = render(request, 'attendance/reports/analytics_pdf.html', {
        'total_students': total_students,
        'total_teachers': total_teachers,
        'total_courses': total_courses,
        'total_sessions': total_sessions,
        'total_attendance': total_attendance,
        'present_count': present_count,
        'absent_count': absent_count,
        'late_count': late_count,
        'overall_present_pct': overall_present_pct,
        'overall_absent_pct': 100 - overall_present_pct,
        'weekly_chart': weekly_chart,
        'dept_chart': dept_chart,
        'college_chart': college_chart,
        'top_courses': top_courses,
        'top_students': top_students,
        'generated_at': timezone.now(),
        'logo_url': _pdf_logo_url(request),
    }).content.decode('utf-8')
    return _pdf_response(html, f'SHAMEL_Analytics_{_date.today()}.pdf', base_url=request.build_absolute_uri('/'))


# ── Medical Excuses ──────────────────────────────────────────────────────────

@login_required
def excuse_portal(request):
    student = get_object_or_404(Student, auth_user=request.user)
    excuses = MedicalExcuse.objects.filter(student=student).order_by('-submitted_at')
    if request.method == 'POST':
        reason = request.POST.get('reason', '').strip()
        schedule_id = request.POST.get('schedule_id') or None
        doc = request.FILES.get('document')
        MedicalExcuse.objects.create(
            student=student, reason=reason,
            schedule_id=schedule_id, document=doc,
        )
        messages.success(request, 'تم تقديم العذر الطبي.')
        return redirect('excuse_portal')
    schedules = Schedule.objects.filter(
        course__in=Enrollment.objects.filter(student=student).values('course')
    ).select_related('course')
    return render(request, 'attendance/excuse_portal.html', {
        'student': student, 'excuses': excuses, 'schedules': schedules,
    })


@login_required
def excuse_approval_board(request):
    is_admin = request.user.is_staff or request.user.is_superuser
    coord = Coordinator.objects.filter(auth_user=request.user).first()
    if not is_admin and not coord:
        return redirect('student_dashboard')
    excuses = MedicalExcuse.objects.select_related('student', 'schedule__course').order_by('-submitted_at')
    # Coordinators see only their college's excuses
    if coord and not is_admin:
        excuses = excuses.filter(student__department__college=coord.college)
    if request.method == 'POST':
        excuse_id = request.POST.get('excuse_id')
        action = request.POST.get('action')
        if action == 'approve':
            action = 'approved'
        elif action == 'reject':
            action = 'rejected'
        note = request.POST.get('review_note') or request.POST.get('reviewer_note') or ''
        excuse = get_object_or_404(MedicalExcuse, pk=excuse_id)
        if action in ('approved', 'rejected'):
            excuse.status = action
            excuse.reviewed_by = request.user
            excuse.review_note = note
            excuse.save()
            # On approval: create/update attendance log so absence is forgiven in %
            if action == 'approved' and excuse.schedule:
                AIAttendanceLog.objects.update_or_create(
                    student=excuse.student,
                    schedule=excuse.schedule,
                    defaults={
                        'status': 'Excused',
                        'timestamp': excuse.submitted_at,
                        'confidence_score': 1.0,
                    },
                )
            messages.success(request, 'تم تحديث حالة العذر.')
        return redirect('excuse_approval_board')
    status_f = request.GET.get('status', '')
    if status_f:
        excuses = excuses.filter(status=status_f)
    return render(request, 'attendance/excuse_board.html', {
        'excuses': excuses, 'status_filter': status_f,
    })


# ── Exam Planner ─────────────────────────────────────────────────────────────

@login_required
@staff_member_required
def exam_planner(request):
    try:
        exams = list(Exam.objects.select_related('course', 'classroom').order_by('date', 'start_time'))
    except Exception:
        exams = []
    if request.method == 'POST':
        action = request.POST.get('action', 'create_exam')
        try:
            if action == 'delete_exam':
                Exam.objects.filter(id=request.POST.get('exam_id')).delete()
                messages.success(request, 'تم حذف الاختبار.')
            elif action == 'assign_seats':
                # seat assignment handled by exam_seating_chart; just acknowledge
                messages.info(request, 'استخدم خريطة المقاعد لتوزيع الطلاب.')
            else:  # create_exam
                # template field is exam_date; accept legacy 'date' too
                exam_date  = request.POST.get('exam_date') or request.POST.get('date')
                course_id  = request.POST.get('course_id')
                start_time = request.POST.get('start_time')
                end_time   = request.POST.get('end_time')
                # All four are NOT NULL in the DB — validate before insert so a
                # missing field gives a clear message instead of a constraint 500.
                missing = []
                if not course_id:  missing.append('المادة')
                if not exam_date:  missing.append('التاريخ')
                if not start_time: missing.append('وقت البداية')
                if not end_time:   missing.append('وقت النهاية')
                if missing:
                    messages.error(request, 'الحقول التالية مطلوبة: ' + '، '.join(missing))
                    return redirect('exam_planner')
                Exam.objects.create(
                    course_id=course_id,
                    exam_type=request.POST.get('exam_type', 'Final'),
                    date=exam_date,
                    start_time=start_time,
                    end_time=end_time,
                    classroom_id=request.POST.get('classroom_id') or None,
                    semester=request.POST.get('semester', ''),
                )
                messages.success(request, 'تمت إضافة الاختبار.')
        except Exception as e:
            messages.error(request, f'خطأ: {e}')
        return redirect('exam_planner')
    return render(request, 'attendance/exam_planner.html', {
        'exams': exams,
        'courses': Course.objects.all(),
        'classrooms': Classroom.objects.all(),
        'semester_choices': SEMESTER_CHOICES_4Y,
    })


@login_required
@staff_member_required
def exam_seating_chart(request):
    exam_id = request.GET.get('exam_id')
    exam = get_object_or_404(Exam, pk=exam_id) if exam_id else None
    seats = ExamSeat.objects.filter(exam=exam).select_related('student') if exam else []
    if request.method == 'POST' and exam:
        student_ids = request.POST.getlist('student_ids')
        for i, sid in enumerate(student_ids, 1):
            ExamSeat.objects.get_or_create(
                exam=exam, student_id=sid,
                defaults={'seat_number': str(i)},
            )
        messages.success(request, 'تم تخصيص المقاعد.')
        return redirect(f'/attendance/exam/seating/?exam_id={exam.pk}')
    exams = Exam.objects.select_related('course').order_by('-date')
    students = Student.objects.all().order_by('name')
    return render(request, 'attendance/exam_seating.html', {
        'exam': exam, 'seats': seats, 'exams': exams, 'students': students,
    })


@login_required
def exam_gate_verify(request):
    """Quick verification at exam gate: check student seat. Staff/gate only."""
    if not (request.user.is_staff or request.user.is_superuser):
        return HttpResponseForbidden('غير مصرح')
    student_code = request.GET.get('code', '').strip()
    exam_id = request.GET.get('exam_id', '')
    seat = None
    student = None
    if student_code and exam_id:
        student = Student.objects.filter(student_code=student_code).first()
        if student:
            seat = ExamSeat.objects.filter(exam_id=exam_id, student=student).first()
    exams = Exam.objects.select_related('course').order_by('-date')
    return render(request, 'attendance/exam_gate_verify.html', {
        'seat': seat, 'student': student, 'exams': exams,
        'exam_id': exam_id, 'student_code': student_code,
    })


# ── Faculty Timeline ─────────────────────────────────────────────────────────

@login_required
@staff_member_required
def faculty_timeline(request):
    teachers = Teacher.objects.select_related('department', 'college').order_by('name')
    teacher_id = request.GET.get('teacher_id')
    data = []
    selected_teacher = None
    if teacher_id:
        try:
            selected_teacher = Teacher.objects.get(pk=teacher_id)
            sessions = LectureSession.objects.filter(
                schedule__teacher=selected_teacher
            ).select_related('schedule__course').order_by('-actual_start_time')[:20]
            data = [{'teacher': selected_teacher, 'sessions': sessions}]
        except Teacher.DoesNotExist:
            pass
    return render(request, 'attendance/faculty_timeline.html', {
        'teachers': teachers,
        'data': data,
        'teacher_id': int(teacher_id) if teacher_id and teacher_id.isdigit() else None,
        'selected_teacher': selected_teacher,
    })


# ── API: student search ──────────────────────────────────────────────────────

@login_required
def api_student_search(request):
    q = request.GET.get('q', '').strip()
    students = Student.objects.filter(
        Q(name__icontains=q) | Q(student_code__icontains=q)
    )[:10] if len(q) >= 2 else []
    data = [{'id': s.pk, 'name': s.name, 'code': s.student_code} for s in students]
    return JsonResponse({'results': data})


# ── System Settings ───────────────────────────────────────────────────────────

@login_required
@staff_member_required
def system_settings(request):
    if request.method == 'POST':
        action = request.POST.get('action', '')

        if action == 'test_email':
            import smtplib, ssl
            host     = request.POST.get('email_host', 'smtp.gmail.com')
            port     = int(request.POST.get('email_port', 587))
            username = request.POST.get('email_user', '')
            password = request.POST.get('email_pass', '')
            try:
                ctx = ssl.create_default_context()
                with smtplib.SMTP(host, port, timeout=8) as server:
                    server.ehlo()
                    server.starttls(context=ctx)
                    if username and password:
                        server.login(username, password)
                return JsonResponse({'ok': True, 'message': 'تم الاتصال بالخادم بنجاح ✓'})
            except smtplib.SMTPAuthenticationError:
                return JsonResponse({'ok': False, 'message': 'بيانات الدخول خاطئة — استخدم App Password لـ Gmail'})
            except smtplib.SMTPConnectError:
                return JsonResponse({'ok': False, 'message': f'تعذر الاتصال بـ {host}:{port}'})
            except Exception as e:
                return JsonResponse({'ok': False, 'message': f'خطأ: {str(e)[:100]}'})

        # Bulk save: key=value pairs from POST
        saved = 0
        for key, value in request.POST.items():
            if key.startswith('cfg_'):
                real_key = key[4:]
                SystemConfig.objects.update_or_create(key=real_key, defaults={'value': value.strip()})
                saved += 1
        if saved:
            SystemConfig.objects.update_or_create(key='setup_done', defaults={'value': '1'})
            messages.success(request, 'تم حفظ الإعدادات.')
        return redirect('system_settings')

    cfg = {c.key: c.value for c in SystemConfig.objects.all()}
    setup_done = SystemConfig.objects.filter(key='setup_done').exists()
    return render(request, 'attendance/system_settings.html', {
        'cfg': cfg,
        'setup_done': setup_done,
    })


# ── Dean Evaluation Dashboard ────────────────────────────────────────────────

@login_required
@staff_member_required
def dean_evaluation_dashboard(request):
    evaluations = CourseEvaluation.objects.select_related('student', 'course') \
                                           .order_by('-submitted_at')
    avg_rating = evaluations.aggregate(avg=Avg('rating'))['avg'] or 0
    course_stats = (evaluations.values('course__title')
                    .annotate(avg=Avg('rating'), count=Count('id'))
                    .order_by('-avg'))
    return render(request, 'attendance/dean_evaluation.html', {
        'evaluations': evaluations[:50],
        'avg_rating': round(avg_rating, 2),
        'course_stats': course_stats,
    })


# ── Project showcase (public landing page) ───────────────────────────────────

def showcase(request):
    from django.conf import settings
    stats = [
        {'value': str(Student.objects.count()),  'label': 'طالب مسجّل'},
        {'value': str(Teacher.objects.count()),  'label': 'عضو هيئة تدريس'},
        {'value': str(Course.objects.count()),   'label': 'مقرر أكاديمي'},
        {'value': str(GateLog.objects.count()),  'label': 'سجل دخول'},
    ]
    features = [
        {'emoji': '🤖', 'title': 'تعرّف على الوجه', 'desc': 'InsightFace buffalo_s — 512-dim cosine، تسجيل 5 زوايا، CLAHE مقاومة الإضاءة الضعيفة'},
        {'emoji': '✅', 'title': 'حضور المحاضرات',  'desc': 'جلسات مباشرة، مسح بالكاميرا أو يدوياً، تتبع الغياب والأعذار الطبية'},
        {'emoji': '🚪', 'title': 'نظام البوابة',    'desc': 'مراقبة دخول الطلاب والأساتذة، كشف الوجه الفوري، سجلات مفصّلة'},
        {'emoji': '📊', 'title': 'تقارير شاملة',    'desc': 'إحصائيات لحظية، تصدير PDF/Excel/CSV، مخططات بيانية تفاعلية'},
        {'emoji': '📡', 'title': 'دعم Offline',     'desc': 'تطبيق Flutter يعمل بدون إنترنت ويتزامن تلقائياً عند الاتصال'},
        {'emoji': '🔔', 'title': 'إشعارات فورية',   'desc': 'WebSocket feed حي، إشعارات الغياب المتكرر، تذكيرات الامتحانات'},
    ]
    roles = [
        {'key': 'admin',       'name': 'المدير العام',     'emoji': '🛡️', 'color': '#EF4444', 'scope': 'نطاق الجامعة كاملاً', 'desc': 'البنية التحتية، سجلات البوابة، دقة الكاميرا، الإحصائيات الكلية، لوحة التدقيق'},
        {'key': 'coordinator', 'name': 'منسق الكلية',     'emoji': '🏫', 'color': '#8B5CF6', 'scope': 'كلية واحدة',          'desc': 'الأعذار المعلّقة، المقررات غير المرصودة، نسبة الحضور بالكلية، الطلاب المعرّضون للخطر'},
        {'key': 'teacher',     'name': 'عضو هيئة التدريس','emoji': '👨‍🏫', 'color': '#3B82F6', 'scope': 'مقرراته فقط',         'desc': 'إدارة الجلسات، تسجيل الحضور بالكاميرا أو يدوياً، الجدول الزمني'},
        {'key': 'student',     'name': 'الطالب',           'emoji': '🎓', 'color': '#10B981', 'scope': 'بياناته الشخصية',     'desc': 'نسبة حضوره، جدوله، الأعذار الطبية، درجاته'},
        {'key': 'gate',        'name': 'موظف البوابة',     'emoji': '🚦', 'color': '#F59E0B', 'scope': 'نقطة الدخول',         'desc': 'سجلات الدخول الفوري، حالة قاعات الدراسة، مسح الوجه عند البوابة'},
    ]
    tech = [
        'Django 4.x', 'Flutter 3.41', 'InsightFace ONNX', 'dlib', 'PostgreSQL',
        'SQLite (offline)', 'Django Channels', 'WebSocket', 'PWA', 'Tailwind CSS',
        'Chart.js', 'MediaPipe', 'sqflite', 'REST API', 'JWT Auth',
    ]
    return render(request, 'attendance/showcase.html', {
        'stats': stats, 'features': features, 'roles': roles, 'tech': tech,
        'debug': settings.DEBUG,
    })


# ── Demo login (DEBUG only) ──────────────────────────────────────────────────

def demo_login(request):
    from django.conf import settings
    if not settings.DEBUG:
        raise Http404
    role = request.GET.get('role', 'admin')
    demo_users = {
        'admin':       'admin',
        'teacher':     'teacher1',
        'student':     'std_demo_1',
        'coordinator': 'coordinator_demo',
        'gate':        'gate',
    }
    username = demo_users.get(role, 'admin')
    user = User.objects.filter(username=username).first()
    if user:
        auth_login(request, user, backend='django.contrib.auth.backends.ModelBackend')
        return _redirect_by_role(request)
    messages.error(request, f'Demo user "{username}" not found.')
    return redirect('login')


# ═══════════════════════════════════════════════════════════════════════════════
# MISSING STUBS (referenced in urls.py but not in helpers)
# ═══════════════════════════════════════════════════════════════════════════════

# These are placeholder implementations that will render a basic page or
# redirect until full templates are available.

# ── Notifications ────────────────────────────────────────────────────────────
# (Full implementations come from _p7_views_addon.py below)


# ═══════════════════════════════════════════════════════════════════════════════
# ▼▼▼  CONTENT FROM HELPER SCRIPTS  ▼▼▼
# ═══════════════════════════════════════════════════════════════════════════════

# ===================================================================
# FROM _views_addon.py  — Attendance Reports / Support Tickets
# ===================================================================

def _build_attendance_queryset(request):
    user = request.user
    is_admin    = user.is_staff or user.is_superuser
    coordinator = Coordinator.objects.filter(auth_user=user).first()
    teacher_obj = Teacher.objects.filter(auth_user=user).first()
    student_obj = Student.objects.filter(auth_user=user).first()

    logs = AIAttendanceLog.objects.select_related(
        'student', 'student__department', 'student__department__college',
        'schedule__course', 'schedule__course__college',
        'schedule__course__department',
    )

    if student_obj:
        logs = logs.filter(student=student_obj)
    elif teacher_obj:
        logs = logs.filter(schedule__teacher=teacher_obj)
    elif coordinator and not is_admin:
        logs = logs.filter(schedule__course__college=coordinator.college)

    f = {
        'college_id':    request.GET.get('college_id', ''),
        'department_id': request.GET.get('department_id', ''),
        'year_level':    request.GET.get('year_level', ''),
        'course_id':     request.GET.get('course_id', ''),
        'semester':      request.GET.get('semester', ''),
        'date_from':     request.GET.get('date_from', ''),
        'date_to':       request.GET.get('date_to', ''),
        'student_id':    request.GET.get('student_id', ''),
        'batch':         request.GET.get('batch', ''),
        'status':        request.GET.get('status', ''),
    }

    if f['college_id']:    logs = logs.filter(schedule__course__college_id=f['college_id'])
    if f['department_id']: logs = logs.filter(schedule__course__department_id=f['department_id'])
    if f['year_level']:    logs = logs.filter(schedule__course__year_level=f['year_level'])
    if f['course_id']:     logs = logs.filter(schedule__course_id=f['course_id'])
    if f['semester']:      logs = logs.filter(schedule__semester=f['semester'])
    if f['date_from']:     logs = logs.filter(timestamp__date__gte=f['date_from'])
    if f['date_to']:       logs = logs.filter(timestamp__date__lte=f['date_to'])
    if f['student_id']:    logs = logs.filter(student_id=f['student_id'])
    if f['batch']:         logs = logs.filter(student__batch__icontains=f['batch'])
    if f['status']:        logs = logs.filter(status=f['status'])

    logs = logs.order_by('student__name', 'timestamp')

    if coordinator and not is_admin:
        colleges    = College.objects.filter(college_id=coordinator.college_id)
        departments = Department.objects.filter(college=coordinator.college)
        courses     = Course.objects.filter(college=coordinator.college)
    else:
        colleges    = College.objects.all()
        departments = Department.objects.filter(college_id=f['college_id']) if f['college_id'] else Department.objects.all()
        courses     = Course.objects.all()

    meta = {
        'colleges': colleges, 'departments': departments, 'courses': courses,
        'year_choices': range(1, 7),
        'semesters': SEMESTER_CHOICES_4Y,
        'is_admin': is_admin, 'coordinator': coordinator,
    }
    return logs, f, meta


def _summarise_logs(logs):
    buckets = defaultdict(lambda: {'present': 0, 'absent': 0, 'excused': 0, 'student': None})
    for log in logs:
        key = log.student_id
        buckets[key]['student'] = log.student
        if log.status in ('Present', 'Late'):
            buckets[key]['present'] += 1
        elif log.status == 'Excused':
            buckets[key]['excused'] += 1  # approved excuses don't count against %
        else:
            buckets[key]['absent'] += 1

    summary = []
    for data in buckets.values():
        total = data['present'] + data['absent']  # excused rows excluded from denominator
        pct   = round(data['present'] / total * 100, 1) if total else 0
        summary.append({
            'student': data['student'],
            'present': data['present'],
            'absent':  data['absent'],
            'excused': data['excused'],
            'total':   total,
            'pct':     pct,
            'flag':    pct < ATTENDANCE_PASS_THRESHOLD,
            'status':  'Pass' if pct >= ATTENDANCE_PASS_THRESHOLD else 'Fail',
        })
    summary.sort(key=lambda x: x['pct'])
    return summary


@login_required
def student_attendance_report(request):
    logs, filters, meta = _build_attendance_queryset(request)
    summary = _summarise_logs(logs)
    below_threshold = sum(1 for s in summary if s['flag'])
    return render(request, 'attendance/reports/student_report.html', {
        'logs': logs, 'summary': summary,
        'below_threshold': below_threshold,
        'filters': filters, **meta,
    })


@login_required
def export_student_attendance_csv(request):
    logs, filters, _ = _build_attendance_queryset(request)
    summary = _summarise_logs(logs)

    response = HttpResponse(content_type='text/csv; charset=utf-8-sig')
    from datetime import date
    college_f = request.GET.get('college', 'All')
    response['Content-Disposition'] = f'attachment; filename="SHAMEL_StudentAttendance_{college_f}_{date.today()}.csv"'
    writer = csv.writer(response)
    writer.writerow([
        'Student ID', 'Student Name', 'College', 'Department',
        'Year Level', 'Batch', 'Present', 'Absent', 'Total Sessions',
        'Attendance %', 'Status (75% threshold)',
    ])
    for row in summary:
        s = row['student']
        dept = getattr(s, 'department', None)
        college = getattr(dept, 'college', None) if dept else None
        writer.writerow([
            getattr(s, 'student_code', s.id),
            s.name,
            college.college_name if college else '',
            dept.name if dept else '',
            getattr(s, 'batch', ''),
            getattr(s, 'semester', ''),
            row['present'], row['absent'], row['total'],
            f"{row['pct']}%",
            row['status'],
        ])
    return response


@login_required
def export_attendance_excel(request):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    logs, filters, meta = _build_attendance_queryset(request)
    summary = _summarise_logs(logs)

    wb = Workbook()
    ws = wb.active
    ws.title = 'Attendance Report'

    BLUE   = 'FF1B263B'
    GREEN  = 'FFD1FAE5'
    YELLOW = 'FFFEF9C3'
    RED    = 'FFFEE2E2'

    headers = ['Student ID', 'Student Name', 'College', 'Department',
               'Year', 'Batch', 'Present', 'Absent', 'Total', 'Attendance %', 'Status']
    ws.append(headers)
    for cell in ws[1]:
        cell.font      = Font(bold=True, color='FFFFFFFF')
        cell.fill      = PatternFill('solid', fgColor=BLUE)
        cell.alignment = Alignment(horizontal='center')

    thin = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'),
    )

    for row in summary:
        s   = row['student']
        dept    = getattr(s, 'department', None)
        college = getattr(dept, 'college', None) if dept else None
        pct     = row['pct']
        color   = GREEN if pct >= 75 else (YELLOW if pct >= 65 else RED)
        data = [
            getattr(s, 'student_code', s.id), s.name,
            college.college_name if college else '',
            dept.name if dept else '',
            getattr(s, 'batch', ''), getattr(s, 'semester', ''),
            row['present'], row['absent'], row['total'],
            f"{pct}%", row['status'],
        ]
        ws.append(data)
        fill = PatternFill('solid', fgColor=color)
        for cell in ws[ws.max_row]:
            cell.fill      = fill
            cell.border    = thin
            cell.alignment = Alignment(horizontal='center')

    for i, col in enumerate(ws.columns, 1):
        ws.column_dimensions[get_column_letter(i)].width = 18

    ws2 = wb.create_sheet('Summary by Course')
    ws2.append(['Course', 'Total Sessions', 'Present', 'Absent', 'Attendance %'])
    for cell in ws2[1]:
        cell.font = Font(bold=True, color='FFFFFFFF')
        cell.fill = PatternFill('solid', fgColor=BLUE)

    course_buckets = defaultdict(lambda: {'present': 0, 'absent': 0})
    for log in logs:
        name = log.schedule.course.title if (log.schedule and log.schedule.course) else 'Unknown'
        if log.status == 'Present':
            course_buckets[name]['present'] += 1
        else:
            course_buckets[name]['absent'] += 1

    for name, data in course_buckets.items():
        total = data['present'] + data['absent']
        avg   = round(data['present'] / total * 100, 1) if total else 0
        ws2.append([name, total, data['present'], data['absent'], f"{avg}%"])

    for i, col in enumerate(ws2.columns, 1):
        ws2.column_dimensions[get_column_letter(i)].width = 22

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    resp = HttpResponse(buf.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    from datetime import date as _date
    resp['Content-Disposition'] = f'attachment; filename="SHAMEL_AttendanceReport_{_date.today()}.xlsx"'
    return resp


@login_required
def teacher_attendance_records(request):
    try:
        teacher = Teacher.objects.get(auth_user=request.user)
    except Teacher.DoesNotExist:
        return _redirect_by_role(request)

    courses  = Course.objects.filter(schedule__teacher=teacher).distinct()
    logs     = AIAttendanceLog.objects.filter(schedule__teacher=teacher).select_related(
        'student', 'schedule__course',
    )

    course_id  = request.GET.get('course_id', '')
    date_from  = request.GET.get('date_from', '')
    date_to    = request.GET.get('date_to', '')
    student_q  = request.GET.get('student_q', '')

    if course_id:   logs = logs.filter(schedule__course_id=course_id)
    if date_from:   logs = logs.filter(timestamp__date__gte=date_from)
    if date_to:     logs = logs.filter(timestamp__date__lte=date_to)
    if student_q:   logs = logs.filter(student__name__icontains=student_q)

    logs = logs.order_by('student__name', '-timestamp')
    summary = _summarise_logs(logs)

    return render(request, 'attendance/teacher_attendance_records.html', {
        'logs': logs, 'summary': summary, 'courses': courses,
        'filters': {'course_id': course_id, 'date_from': date_from, 'date_to': date_to, 'student_q': student_q},
        'below_threshold': sum(1 for s in summary if s['flag']),
    })


@login_required
def tickets_list(request):
    user     = request.user
    is_admin = user.is_staff or user.is_superuser
    coord    = Coordinator.objects.filter(auth_user=user).first()

    if is_admin:
        tickets = SupportTicket.objects.all()
    else:
        tickets = SupportTicket.objects.filter(user=user)

    status_filter = request.GET.get('status', '')
    if status_filter:
        tickets = tickets.filter(status=status_filter)

    tickets = tickets.order_by('-created_at').select_related('user')

    return render(request, 'attendance/tickets_list.html', {
        'tickets': tickets, 'is_admin': is_admin,
        'status_choices': SupportTicket.STATUS_CHOICES,
        'selected_status': status_filter,
    })


@login_required
def ticket_detail(request, ticket_id):
    user     = request.user
    is_admin = user.is_staff or user.is_superuser
    ticket   = get_object_or_404(SupportTicket, id=ticket_id)

    can_view = (ticket.user == user or is_admin)
    if not can_view:
        messages.error(request, 'غير مصرح بعرض هذا البلاغ.')
        return redirect('tickets_list')

    if request.method == 'POST' and is_admin:
        reply = request.POST.get('admin_reply', '').strip()
        new_status = request.POST.get('status', ticket.status)
        ticket.admin_reply = reply
        ticket.status = new_status
        ticket.save()
        messages.success(request, 'تم تحديث البلاغ.')
        return redirect('ticket_detail', ticket_id=ticket_id)

    return render(request, 'attendance/ticket_detail.html', {
        'ticket': ticket, 'is_admin': is_admin,
    })


@login_required
def create_ticket(request):
    user = request.user

    if request.method == 'POST':
        subject  = request.POST.get('subject', '').strip()
        # form sends 'description', fallback to 'body' for compatibility
        body     = (request.POST.get('description') or request.POST.get('body', '')).strip()
        priority = request.POST.get('priority', 'medium')

        ticket_type = request.POST.get('ticket_type', 'general')
        if not subject or not body:
            messages.error(request, 'الموضوع والوصف مطلوبان.')
        else:
            # Dual-DB aware: the VPS PostgreSQL schema has extra NOT NULL columns
            # (description/requester_id/ticket_type) not present in the Django model
            # nor in the local SQLite fallback. Branch on vendor so BOTH work.
            from django.db import connection as _conn
            ticket = None
            try:
                if _conn.vendor == 'postgresql':
                    with _conn.cursor() as cur:
                        cur.execute(
                            """INSERT INTO attendance_supportticket
                               (subject, description, body, status, priority, ticket_type,
                                created_at, updated_at, requester_id, user_id, admin_reply)
                               VALUES (%s, %s, %s, 'open', %s, %s, NOW(), NOW(), %s, %s, '')
                               RETURNING id""",
                            [subject, body, body, priority, ticket_type, user.id, user.id]
                        )
                        ticket_id = cur.fetchone()[0]
                    ticket = SupportTicket.objects.get(id=ticket_id)
                else:
                    # SQLite / other — plain ORM works against the model fields
                    ticket = SupportTicket.objects.create(
                        user=user, subject=subject, body=body,
                        status='open', priority=priority,
                    )
            except Exception as e:
                messages.error(request, f'خطأ في حفظ البلاغ: {e}')
                ticket = None
            if ticket:
                notif_body = f'{user.get_full_name() or user.username} رفع بلاغاً: {subject}'
                notif_title = f'بلاغ جديد #{ticket.id}'
                # Notify all admin users — bulk_create (one query)
                admin_users = list(User.objects.filter(is_staff=True).values_list('id', flat=True))
                Notification.objects.bulk_create([
                    Notification(user_id=uid, title=notif_title, body=notif_body, level='warning')
                    for uid in admin_users
                ], ignore_conflicts=True)
                # Notify coordinators scoped to submitter's college — bulk_create
                try:
                    college = None
                    st = Student.objects.filter(auth_user=user).select_related('department__college').first()
                    if st and st.department:
                        college = st.department.college
                    if college is None:
                        tc = Teacher.objects.filter(auth_user=user).first()
                        if tc:
                            college = getattr(tc, 'college', None)
                    if college:
                        coord_user_ids = list(
                            Coordinator.objects.filter(college=college, auth_user__isnull=False)
                            .values_list('auth_user_id', flat=True)
                        )
                        Notification.objects.bulk_create([
                            Notification(user_id=uid, title=notif_title, body=notif_body, level='warning')
                            for uid in coord_user_ids
                        ], ignore_conflicts=True)
                except Exception:
                    pass
                messages.success(request, f'تم رفع البلاغ #{ticket.id} بنجاح.')
                return redirect('ticket_detail', ticket_id=ticket.id)
            return redirect('tickets_list')

    TYPE_CHOICES = [
        ('general', 'عام'),
        ('technical', 'تقني'),
        ('academic', 'أكاديمي'),
        ('other', 'أخرى'),
    ]
    return render(request, 'attendance/create_ticket.html', {
        'priority_choices': SupportTicket.PRIORITY_CHOICES,
        'type_choices': TYPE_CHOICES,
    })


@login_required
def admin_tickets(request):
    if not (request.user.is_staff or request.user.is_superuser):
        return redirect('tickets_list')

    tickets = SupportTicket.objects.all().order_by('-created_at').select_related('user')

    status_f   = request.GET.get('status', '')
    priority_f = request.GET.get('priority', '')

    if status_f:   tickets = tickets.filter(status=status_f)
    if priority_f: tickets = tickets.filter(priority=priority_f)

    stats = {
        'open':        SupportTicket.objects.filter(status='open').count(),
        'in_progress': SupportTicket.objects.filter(status='in_progress').count(),
        'closed':      SupportTicket.objects.filter(status='closed').count(),
    }

    return render(request, 'attendance/admin_tickets.html', {
        'tickets': tickets, 'stats': stats,
        'status_choices':    SupportTicket.STATUS_CHOICES,
        'priority_choices':  SupportTicket.PRIORITY_CHOICES,
        'selected_status':   status_f,
        'selected_priority': priority_f,
    })


# ===================================================================
# FROM _missing_views.py  — Search, Edit CRUD, Detail pages
# ===================================================================

@login_required
def export_search_pdf(request):
    """Export global search results as a professional PDF."""
    from datetime import date as _date
    query = request.GET.get('q', '').strip()
    results = {'students': [], 'teachers': [], 'courses': [], 'classrooms': []}
    if len(query) >= 2:
        results['students']   = list(Student.objects.filter(name__icontains=query).select_related('department')[:30])
        results['teachers']   = list(Teacher.objects.filter(name__icontains=query).select_related('department', 'college')[:30])
        results['courses']    = list(Course.objects.filter(Q(title__icontains=query) | Q(course_code__icontains=query)).select_related('department', 'college')[:20])
        results['classrooms'] = list(Classroom.objects.filter(name__icontains=query)[:20])
    total = sum(len(v) for v in results.values())
    html = render(request, 'attendance/reports/search_pdf.html', {
        'query': query, 'results': results, 'total': total,
        'generated_at': timezone.now(),
        'logo_url': _pdf_logo_url(request),
    }).content.decode('utf-8')
    safe_q = query.replace(' ', '_')[:20] if query else 'all'
    return _pdf_response(html, f'SHAMEL_Search_{safe_q}_{_date.today()}.pdf', base_url=request.build_absolute_uri('/'))


@login_required
def global_search(request):
    query = request.GET.get('q', '').strip()
    results = {
        'students': [], 'teachers': [], 'courses': [],
        'classrooms': [], 'tickets': [],
    }
    total = 0
    if len(query) >= 2:
        results['students']   = Student.objects.filter(name__icontains=query).select_related('department')[:8]
        results['teachers']   = Teacher.objects.filter(name__icontains=query).select_related('department')[:8]
        results['courses']    = Course.objects.filter(
            Q(title__icontains=query) | Q(course_code__icontains=query)
        )[:8]
        results['classrooms'] = Classroom.objects.filter(name__icontains=query)[:6]
        if request.user.is_staff or request.user.is_superuser:
            results['tickets'] = SupportTicket.objects.filter(
                Q(subject__icontains=query) | Q(body__icontains=query)
            )[:6]
        total = sum(len(v) for v in results.values())

    if request.headers.get('Accept') == 'application/json':
        data = []
        for s in results['students']:
            data.append({'title': s.name, 'sub': 'طالب', 'icon': 'person',
                         'url': f'/attendance/students/{s.id}/'})
        for t in results['teachers']:
            data.append({'title': t.name, 'sub': 'أستاذ', 'icon': 'school',
                         'url': f'/attendance/teachers/{t.teacher_id}/'})
        for c in results['courses']:
            data.append({'title': c.title, 'sub': c.course_code, 'icon': 'menu_book',
                         'url': '/attendance/courses/'})
        return JsonResponse({'results': data})

    return render(request, 'attendance/search_results.html', {
        'query': query, 'results': results, 'total': total,
    })


@login_required
def edit_course(request, course_id):
    if not (request.user.is_staff or Coordinator.objects.filter(auth_user=request.user).exists()):
        messages.error(request, 'غير مصرح.')
        return redirect('courses_list')

    course = get_object_or_404(Course, pk=course_id)
    coordinator = Coordinator.objects.filter(auth_user=request.user).first()
    is_admin = request.user.is_staff or request.user.is_superuser

    if request.method == 'POST':
        course.course_code = request.POST.get('course_code', course.course_code).strip()
        course.title       = request.POST.get('title', course.title).strip()
        course.credits     = int(request.POST.get('credits', course.credits))
        course.total_hours = int(request.POST.get('total_hours', course.total_hours))
        course.year_level  = request.POST.get('year_level') or course.year_level
        if is_admin:
            college_id = request.POST.get('college_id')
            dept_id    = request.POST.get('department_id')
            if college_id:
                course.college    = College.objects.filter(pk=college_id).first()
            if dept_id:
                course.department = Department.objects.filter(pk=dept_id).first()
        course.save()
        messages.success(request, 'تم تحديث بيانات المادة بنجاح.')
        return redirect('courses_list')

    colleges    = College.objects.all() if is_admin else College.objects.filter(pk=coordinator.college_id)
    departments = Department.objects.filter(college=course.college) if course.college else Department.objects.all()
    return render(request, 'attendance/edit_course.html', {
        'course': course, 'colleges': colleges, 'departments': departments,
        'year_choices': range(1, 7), 'is_admin': is_admin,
    })


@login_required
def edit_classroom(request, classroom_id):
    if not request.user.is_staff:
        messages.error(request, 'غير مصرح.')
        return redirect('classrooms_list')
    classroom = get_object_or_404(Classroom, pk=classroom_id)
    if request.method == 'POST':
        classroom.name           = request.POST.get('name', classroom.name).strip()
        classroom.location       = request.POST.get('location', '').strip()
        classroom.capacity       = int(request.POST.get('capacity', classroom.capacity))
        classroom.classroom_type = request.POST.get('classroom_type', classroom.classroom_type)
        classroom.save()
        messages.success(request, 'تم تحديث بيانات القاعة.')
        return redirect('classrooms_list')
    return render(request, 'attendance/edit_classroom.html', {
        'classroom': classroom,
        'types': Classroom.CLASSROOM_TYPES,
    })


@login_required
def edit_schedule(request, schedule_id):
    coordinator = Coordinator.objects.filter(auth_user=request.user).first()
    is_admin    = request.user.is_staff or request.user.is_superuser
    if not (is_admin or coordinator):
        messages.error(request, 'غير مصرح.')
        return redirect('schedule')

    try:
        schedule = Schedule.objects.get(pk=schedule_id)
    except Schedule.DoesNotExist:
        messages.error(request, f'الجدول رقم {schedule_id} غير موجود — تأكد من الاتصال بالسيرفر الرئيسي.')
        return redirect('schedule')

    if request.method == 'POST':
        schedule.day_of_week = request.POST.get('day_of_week', schedule.day_of_week)
        schedule.start_time  = request.POST.get('start_time', schedule.start_time)
        schedule.end_time    = request.POST.get('end_time', schedule.end_time)
        schedule.batch       = request.POST.get('batch', schedule.batch)
        schedule.semester    = request.POST.get('semester', schedule.semester)
        course_id    = request.POST.get('course_id')
        teacher_id   = request.POST.get('teacher_id')
        classroom_id = request.POST.get('classroom_id')
        if course_id:    schedule.course    = get_object_or_404(Course, pk=course_id)
        if teacher_id:   schedule.teacher   = get_object_or_404(Teacher, pk=teacher_id)
        if classroom_id: schedule.classroom = get_object_or_404(Classroom, pk=classroom_id)
        schedule.save()
        messages.success(request, 'تم تحديث المحاضرة.')
        return redirect('schedule')

    if coordinator and not is_admin:
        courses    = Course.objects.filter(college=coordinator.college)
        teachers   = Teacher.objects.filter(college=coordinator.college)
        classrooms = Classroom.objects.all()
    else:
        courses    = Course.objects.all()
        teachers   = Teacher.objects.all()
        classrooms = Classroom.objects.all()

    days = ['Saturday', 'Sunday', 'Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday']
    return render(request, 'attendance/edit_schedule.html', {
        'schedule': schedule, 'courses': courses, 'teachers': teachers,
        'classrooms': classrooms, 'days': days, 'semesters': SEMESTER_CHOICES_4Y,
        'coordinator': coordinator, 'is_admin': is_admin,
    })


@login_required
def student_detail(request, student_id):
    user        = request.user
    is_admin    = user.is_staff or user.is_superuser
    coordinator = Coordinator.objects.filter(auth_user=user).first()
    teacher_obj = Teacher.objects.filter(auth_user=user).first()

    student = get_object_or_404(Student, pk=student_id)

    if not is_admin and not coordinator and not teacher_obj:
        own = Student.objects.filter(auth_user=user).first()
        if not own or own.pk != student.pk:
            messages.error(request, 'غير مصرح.')
            return redirect('student_dashboard')
    if coordinator and not is_admin:
        if not student.department or student.department.college != coordinator.college:
            messages.error(request, 'هذا الطالب ليس في كليتك.')
            return redirect('coordinator_students')

    logs = AIAttendanceLog.objects.filter(student=student).select_related('schedule__course').order_by('-timestamp')
    total   = logs.count()
    present = logs.filter(status='Present').count()
    pct     = round(present / total * 100, 1) if total else 0

    course_stats = defaultdict(lambda: {'present': 0, 'total': 0, 'name': ''})
    for log in logs:
        name = log.schedule.course.title if (log.schedule and log.schedule.course) else 'غير محدد'
        course_stats[name]['name']  = name
        course_stats[name]['total'] += 1
        if log.status == 'Present':
            course_stats[name]['present'] += 1

    course_summary = []
    for cname, data in course_stats.items():
        cpct = round(data['present'] / data['total'] * 100, 1) if data['total'] else 0
        course_summary.append({
            'name': cname, 'present': data['present'],
            'total': data['total'], 'pct': cpct, 'flag': cpct < 75,
        })
    course_summary.sort(key=lambda x: x['pct'])

    fin     = FinancialStatus.objects.filter(student=student).first()
    enrolls = Enrollment.objects.filter(student=student).select_related('course', 'classroom')
    absent  = total - present

    # info_items used by template info card
    dept = student.department
    info_items = [
        ('رمز الطالب', student.student_code),
        ('الاسم', student.name),
        ('القسم', dept.name if dept else '—'),
        ('الكلية', dept.college.college_name if (dept and dept.college) else '—'),
    ]

    return render(request, 'attendance/student_detail.html', {
        'student':        student,
        'logs':           logs[:20],
        'total':          total,
        'present':        present,
        'absent':         absent,
        'pct':            pct,
        'course_summary': course_summary,
        'fin':            fin,
        'enrolls':        enrolls,
        'info_items':     info_items,
        'is_admin':       is_admin,
        'coordinator':    coordinator,
    })


@login_required
def teacher_detail(request, teacher_id):
    user        = request.user
    is_admin    = user.is_staff or user.is_superuser
    coordinator = Coordinator.objects.filter(auth_user=user).first()
    if not (is_admin or coordinator):
        messages.error(request, 'غير مصرح.')
        return redirect('professor_dashboard')

    teacher = get_object_or_404(Teacher, pk=teacher_id)
    if coordinator and not is_admin:
        if teacher.college != coordinator.college:
            messages.error(request, 'هذا الأستاذ ليس في كليتك.')
            return redirect('coordinator_faculty')

    schedules   = Schedule.objects.filter(teacher=teacher).select_related('course', 'classroom')
    total_students_served = AIAttendanceLog.objects.filter(
        schedule__teacher=teacher
    ).values('student').distinct().count()
    total_sessions = AIAttendanceLog.objects.filter(schedule__teacher=teacher).count()

    return render(request, 'attendance/teacher_detail.html', {
        'teacher': teacher, 'schedules': schedules,
        'total_students_served': total_students_served,
        'total_sessions': total_sessions,
        'is_admin': is_admin, 'coordinator': coordinator,
    })


@login_required
def audit_log_view(request):
    if not request.user.is_staff:
        return redirect('admin_panel')

    try:
        logs = AuditLog.objects.select_related('user').order_by('-timestamp')

        action_f = request.GET.get('action', '')
        model_f  = request.GET.get('model', '')
        user_f   = request.GET.get('user', '')
        date_f   = request.GET.get('date', '')

        if action_f: logs = logs.filter(action=action_f)
        if model_f:  logs = logs.filter(target_model__icontains=model_f)
        if user_f:   logs = logs.filter(user__username__icontains=user_f)
        if date_f:
            try:
                from datetime import date as _date
                _date.fromisoformat(date_f)
                logs = logs.filter(timestamp__date=date_f)
            except (ValueError, TypeError):
                pass

        logs = list(logs[:500])
    except Exception as e:
        logs = []
        action_f = model_f = user_f = date_f = ''

    action_choices = [
        ('CREATE', 'إنشاء'), ('UPDATE', 'تعديل'), ('DELETE', 'حذف'),
        ('LOGIN', 'تسجيل دخول'), ('LOGOUT', 'تسجيل خروج'),
        ('TOGGLE_ACCESS', 'تغيير صلاحية'), ('TOGGLE_STUDENT_ACCESS', 'تغيير صلاحية طالب'),
        ('REGISTER_STUDENT', 'تسجيل طالب'), ('REGISTER_TEACHER', 'تسجيل أستاذ'),
        ('OPEN_SESSION', 'فتح جلسة'), ('STOP_SESSION', 'إغلاق جلسة'),
        ('ENROLL_FACE', 'تسجيل وجه'),
    ]
    return render(request, 'attendance/audit_log.html', {
        'logs': logs,
        'actions': action_choices,
        'filters': {'action': action_f, 'model': model_f, 'user': user_f, 'date': date_f},
    })


@login_required
def departments_view(request):
    if not request.user.is_staff:
        return redirect('admin_panel')

    if request.method == 'POST':
        action = request.POST.get('action', '')

        if action == 'add_college':
            name = request.POST.get('college_name', '').strip()
            if name:
                College.objects.create(college_name=name, name=name)
                messages.success(request, 'تمت إضافة الكلية.')

        elif action == 'delete_college':
            cid = request.POST.get('college_id')
            College.objects.filter(pk=cid).delete()
            messages.success(request, 'تم حذف الكلية.')

        elif action == 'add_department':
            name = request.POST.get('dept_name', '').strip()
            college_id = request.POST.get('college_id')
            if name and college_id:
                college = College.objects.filter(pk=college_id).first()
                if college:
                    Department.objects.create(name=name, college=college)
                    messages.success(request, 'تمت إضافة القسم.')

        elif action == 'delete_department':
            did = request.POST.get('dept_id')
            Department.objects.filter(pk=did).delete()
            messages.success(request, 'تم حذف القسم.')

        elif action == 'edit_department':
            did  = request.POST.get('dept_id')
            name = request.POST.get('dept_name', '').strip()
            if did and name:
                Department.objects.filter(pk=did).update(name=name)
                messages.success(request, 'تم تعديل القسم.')

        elif action == 'edit_college':
            cid  = request.POST.get('college_id')
            name = request.POST.get('college_name', '').strip()
            if cid and name:
                College.objects.filter(pk=cid).update(college_name=name, name=name)
                messages.success(request, 'تم تعديل الكلية.')

        return redirect('departments_view')

    colleges    = College.objects.prefetch_related('department_set').all()
    departments = Department.objects.select_related('college').all()
    return render(request, 'attendance/departments.html', {
        'colleges': colleges, 'departments': departments,
    })


@login_required
def student_schedule_view(request):
    student = get_object_or_404(Student, auth_user=request.user)
    enrollments = Enrollment.objects.filter(student=student).select_related('course', 'classroom')
    enrolled_courses = [e.course for e in enrollments]
    schedules = Schedule.objects.filter(
        course__in=enrolled_courses
    ).select_related('course', 'teacher', 'classroom').order_by('day_of_week', 'start_time')

    DAYS_ORDER = ['Saturday', 'Sunday', 'Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday']
    DAY_AR = {
        'Saturday': 'السبت', 'Sunday': 'الأحد', 'Monday': 'الاثنين',
        'Tuesday': 'الثلاثاء', 'Wednesday': 'الأربعاء',
        'Thursday': 'الخميس', 'Friday': 'الجمعة',
    }
    schedule_by_day = {}
    for day in DAYS_ORDER:
        day_scheds = [s for s in schedules if s.day_of_week == day]
        if day_scheds:
            schedule_by_day[day] = {'ar': DAY_AR[day], 'schedules': day_scheds}

    return render(request, 'attendance/student_schedule.html', {
        'student': student, 'schedule_by_day': schedule_by_day,
        'total': schedules.count(),
    })


@login_required
def edit_student(request, student_id):
    user        = request.user
    is_admin    = user.is_staff or user.is_superuser
    coordinator = Coordinator.objects.filter(auth_user=user).first()
    if not (is_admin or coordinator):
        messages.error(request, 'غير مصرح.')
        return redirect('student_dashboard')

    student = get_object_or_404(Student, pk=student_id)
    if coordinator and not is_admin:
        if not student.department or student.department.college != coordinator.college:
            messages.error(request, 'ليس في نطاق كليتك.')
            return redirect('coordinator_students')

    BLOOD_TYPES = ['A+','A-','B+','B-','AB+','AB-','O+','O-']
    if request.method == 'POST':
        student.name             = request.POST.get('name', student.name).strip()
        student.phone_number     = request.POST.get('phone_number', '').strip() or ''
        student.university_email = request.POST.get('university_email', '').strip() or ''
        student.batch            = request.POST.get('batch', student.batch)
        student.is_registered    = 'is_registered' in request.POST
        student.is_allowed_entry = 'is_allowed_entry' in request.POST
        dept_id = request.POST.get('department_id')
        if dept_id:
            student.department = Department.objects.filter(pk=dept_id).first()
        if 'profile_photo' in request.FILES:
            student.face_image = request.FILES['profile_photo']
        student.save()
        # nationality and blood_type are in DB but not model — raw SQL
        nat_choice  = request.POST.get('nationality_choice', 'Sudan')
        nationality = 'Sudan' if nat_choice == 'Sudan' else request.POST.get('nationality_text', '').strip()
        blood_type  = request.POST.get('blood_type', '').strip()
        # nationality/blood_type exist on the VPS schema but not on the local
        # SQLite fallback — degrade gracefully instead of 500.
        from django.db import connection as _c, OperationalError, ProgrammingError
        try:
            _c.cursor().execute(
                'UPDATE attendance_student SET nationality=%s, blood_type=%s WHERE id=%s',
                [nationality or None, blood_type or None, student.pk]
            )
        except (OperationalError, ProgrammingError):
            pass
        if student.auth_user:
            student.auth_user.email = student.university_email or student.auth_user.email
            student.auth_user.save(update_fields=['email'])
        messages.success(request, 'تم تحديث بيانات الطالب.')
        return redirect('student_detail', student_id=student.pk)

    departments = (Department.objects.filter(college=coordinator.college)
                   if (coordinator and not is_admin)
                   else Department.objects.select_related('college').all())
    colleges = College.objects.all() if is_admin else College.objects.none()
    # Fetch nationality/blood_type from DB (VPS-only columns — safe fallback)
    from django.db import connection as _c, OperationalError, ProgrammingError
    student.nationality = ''
    student.blood_type  = ''
    try:
        cur = _c.cursor()
        cur.execute('SELECT nationality, blood_type FROM attendance_student WHERE id=%s', [student.pk])
        row = cur.fetchone()
        if row:
            student.nationality = row[0] or ''
            student.blood_type  = row[1] or ''
    except (OperationalError, ProgrammingError):
        pass
    return render(request, 'attendance/edit_student.html', {
        'student': student, 'departments': departments, 'colleges': colleges,
        'is_admin': is_admin, 'coordinator': coordinator,
        'blood_types': BLOOD_TYPES,
        'semesters': SEMESTER_CHOICES_4Y,
    })


@login_required
def edit_teacher(request, teacher_id):
    user        = request.user
    is_admin    = user.is_staff or user.is_superuser
    coordinator = Coordinator.objects.filter(auth_user=user).first()
    if not (is_admin or coordinator):
        messages.error(request, 'غير مصرح.')
        return redirect('professor_dashboard')

    teacher = get_object_or_404(Teacher, pk=teacher_id)
    if coordinator and not is_admin:
        if teacher.college != coordinator.college:
            messages.error(request, 'ليس في نطاق كليتك.')
            return redirect('coordinator_faculty')

    BLOOD_TYPES = ['A+','A-','B+','B-','AB+','AB-','O+','O-']
    if request.method == 'POST':
        teacher.name               = request.POST.get('name', teacher.name).strip()
        teacher.academic_degree    = request.POST.get('academic_degree', teacher.academic_degree)
        teacher.major              = request.POST.get('major', teacher.major).strip()
        teacher.phone_number       = request.POST.get('phone_number', '').strip() or ''
        teacher.university_email   = request.POST.get('university_email', '').strip() or ''
        # is_allowed_entry for teachers must NOT be toggled via UI — gate access is implicit for staff
        if request.POST.get('gender'):
            teacher.gender = request.POST.get('gender')
        dept_id = request.POST.get('department_id')
        if dept_id:
            teacher.department = Department.objects.filter(pk=dept_id).first()
        if is_admin:
            college_id = request.POST.get('college_id')
            if college_id:
                teacher.college = College.objects.filter(pk=college_id).first()
        if 'profile_photo' in request.FILES:
            teacher.face_image = request.FILES['profile_photo']
        teacher.save()
        # blood_type and nationality in DB but not model
        blood_type  = request.POST.get('blood_type', '').strip()
        nat_choice  = request.POST.get('nationality_choice', 'Sudan')
        nationality = 'Sudan' if nat_choice == 'Sudan' else request.POST.get('nationality_text', '').strip()
        from django.db import connection as _c, OperationalError, ProgrammingError
        try:
            _c.cursor().execute(
                'UPDATE attendance_teacher SET blood_type=%s WHERE teacher_id=%s',
                [blood_type or None, teacher.pk]
            )
        except (OperationalError, ProgrammingError):
            pass
        messages.success(request, 'تم تحديث بيانات الأستاذ.')
        return redirect('teacher_detail', teacher_id=teacher.pk)

    colleges    = College.objects.all() if is_admin else College.objects.filter(pk=coordinator.college_id)
    departments = (Department.objects.filter(college=coordinator.college)
                   if (coordinator and not is_admin)
                   else Department.objects.select_related('college').all())
    # Fetch blood_type from DB (VPS-only column — safe fallback)
    from django.db import connection as _c, OperationalError, ProgrammingError
    teacher.blood_type = ''
    try:
        cur = _c.cursor()
        cur.execute('SELECT blood_type FROM attendance_teacher WHERE teacher_id=%s', [teacher.pk])
        row = cur.fetchone()
        teacher.blood_type = (row[0] or '') if row else ''
    except (OperationalError, ProgrammingError):
        pass
    teacher.nationality = 'Sudan'  # teachers default to Sudan
    return render(request, 'attendance/edit_teacher.html', {
        'teacher': teacher, 'colleges': colleges, 'departments': departments,
        'is_admin': is_admin, 'coordinator': coordinator,
        'blood_types': BLOOD_TYPES,
    })


# ===================================================================
# FROM phase6_views_append.py
# ===================================================================

def face_login(request):
    if request.method == 'GET':
        return render(request, 'attendance/face_login.html')

    if request.method == 'POST':
        b64_data = request.POST.get('image_data', '')
        if not b64_data:
            return JsonResponse({'status': 'error', 'message': 'No image data'}, status=400)

        try:
            if ',' in b64_data:
                b64_data = b64_data.split(',')[1]
            # Fix base64 padding and whitespace issues
            b64_data = b64_data.strip().replace(' ', '+')
            padding = 4 - len(b64_data) % 4
            if padding != 4:
                b64_data += '=' * padding
            try:
                img_bytes = base64.b64decode(b64_data)
            except Exception:
                return JsonResponse({'status': 'error', 'message': 'بيانات الصورة غير صالحة — أعد المحاولة'}, status=400)

            matched_name = None
            matched_type = None

            live_encoding = None
            try:
                if FACE_ENGINE_AVAILABLE and NUMPY_AVAILABLE and CV2_AVAILABLE:
                    img_arr = cv2.imdecode(
                        np.frombuffer(img_bytes, dtype=np.uint8), cv2.IMREAD_COLOR
                    )
                    if img_arr is not None:
                        enc = _fe.encode(img_arr[:, :, ::-1])  # BGR→RGB
                        live_encoding = enc
            except Exception:
                return JsonResponse({'status': 'error', 'message': 'تعذر تحليل الصورة — تأكد من وضوح الإضاءة وأن الوجه ظاهر بوضوح'})

            if live_encoding is None:
                return JsonResponse({'status': 'error', 'message': 'لم يتم اكتشاف وجه — ضع وجهك أمام الكاميرا مباشرة'})

            matched_name = matched_type = matched_pk = None
            if NUMPY_AVAILABLE and live_encoding:
                matched_name, matched_type, matched_pk = match_face_from_db(live_encoding)

            if not matched_name:
                return JsonResponse({
                    'status': 'error',
                    'code': 'face_not_registered',
                    'message': 'الوجه غير مسجل في النظام — يرجى التسجيل أولاً أو استخدام كلمة المرور'
                })

            auth_user = None
            person_obj = None

            # Use matched_pk from match_face_from_db — exact PK lookup, no icontains
            if matched_type == 'student':
                student = Student.objects.filter(pk=matched_pk).select_related('auth_user').first()
                teacher = None
            else:
                student = None
                teacher = Teacher.objects.filter(pk=matched_pk).select_related('auth_user').first()

            if student:
                person_obj = student
                auth_user = student.auth_user  # must be pre-linked by admin — no auto-creation
            elif teacher:
                person_obj = teacher
                auth_user = teacher.auth_user

            if not auth_user:
                # Account not pre-linked — admin must link via user management UI
                return JsonResponse({
                    'status': 'error',
                    'message': 'الوجه معروف لكن لا يوجد حساب مرتبط — راجع الإدارة لربط حسابك'
                })

            auth_login(request, auth_user, backend='django.contrib.auth.backends.ModelBackend')

            if auth_user.is_superuser or auth_user.is_staff:
                redirect_url = '/attendance/admin-panel/'
            elif Coordinator.objects.filter(auth_user=auth_user).exists():
                redirect_url = '/attendance/coordinator/dashboard/'
            elif Teacher.objects.filter(auth_user=auth_user).exists():
                redirect_url = '/attendance/professor-dashboard/'
            elif Student.objects.filter(auth_user=auth_user).exists():
                redirect_url = '/attendance/student/dashboard/'
            else:
                redirect_url = '/attendance/gate/'

            return JsonResponse({'status': 'success', 'redirect': redirect_url, 'name': matched_name})

        except Exception as e:
            import logging
            logging.getLogger('attendance').error('face_login error: %s', e, exc_info=True)
            return JsonResponse({'status': 'error', 'message': 'حدث خطأ غير متوقع — أعد المحاولة أو استخدم كلمة المرور'})

    return JsonResponse({'status': 'error'}, status=400)


@login_required
def teacher_profile_view(request):
    try:
        teacher = Teacher.objects.get(auth_user=request.user)
    except Teacher.DoesNotExist:
        return _redirect_by_role(request)
    schedules = Schedule.objects.filter(teacher=teacher).select_related(
        'course', 'classroom').order_by('day_of_week', 'start_time')
    sessions_completed = LectureSession.objects.filter(
        schedule__teacher=teacher, is_active=False).count()
    try:
        permissions = list(ClassroomPermission.objects.filter(teacher=teacher).select_related('classroom'))
    except (AttributeError, TypeError):
        permissions = []
    return render(request, 'attendance/teacher_profile.html', {
        'teacher': teacher,
        'schedules': schedules,
        'sessions_completed': sessions_completed,
        'permissions': permissions,
    })


@login_required
def classrooms_status_view(request):
    try:
        from django.db import close_old_connections
        close_old_connections()
        now = timezone.now()
        classrooms = Classroom.objects.all().order_by('name')
        classroom_data = []
        for room in classrooms:
            current_sched = Schedule.objects.filter(
                classroom=room,
                day_of_week=now.strftime('%A'),
                start_time__lte=now.time(),
                end_time__gte=now.time()
            ).select_related('course', 'teacher').first()
            classroom_data.append({'room': room, 'current_schedule': current_sched})
        return render(request, 'attendance/classrooms_status.html', {'classroom_data': classroom_data, 'db_error': False})
    except Exception as e:
        return render(request, 'attendance/classrooms_status.html', {
            'classroom_data': [],
            'db_error': True,
            'db_error_msg': 'تعذّر الاتصال بقاعدة البيانات. تحقق من الاتصال بالشبكة.',
        })


@login_required
def teacher_permissions_view(request):
    teacher = get_object_or_404(Teacher, auth_user=request.user)
    try:
        permissions = list(ClassroomPermission.objects.filter(
            teacher=teacher).select_related('classroom').order_by('classroom__name'))
    except (AttributeError, TypeError):
        permissions = []
    return render(request, 'attendance/teacher_permissions.html', {
        'permissions': permissions, 'teacher': teacher,
    })


@login_required
@staff_member_required
def gate_reports(request):
    date_filter  = request.GET.get('date')
    type_filter  = request.GET.get('user_type', '')
    status_filter = request.GET.get('status', '')
    today = timezone.now().date()
    try:
        qs = GateLog.objects.select_related('student', 'teacher').order_by('-timestamp')
        if date_filter:
            qs = qs.filter(timestamp__date=date_filter)
        if status_filter:
            qs = qs.filter(status=status_filter)
        if type_filter == 'Student':
            qs = qs.filter(student__isnull=False)
        elif type_filter == 'Teacher':
            qs = qs.filter(teacher__isnull=False)

        today_qs      = GateLog.objects.filter(timestamp__date=today)
        total_today   = today_qs.count()
        allowed_today = today_qs.filter(status='Allowed').count()
        denied_today  = today_qs.filter(status='Denied').count()

        by_type = [
            {'user_type': 'Student', 'count': today_qs.filter(student__isnull=False).count()},
            {'user_type': 'Teacher', 'count': today_qs.filter(teacher__isnull=False).count()},
        ]

        logs = list(qs[:200])
    except Exception:
        logs = []; total_today = 0; allowed_today = 0; denied_today = 0; by_type = []

    return render(request, 'attendance/gate_reports.html', {
        'logs':           logs,
        'total_today':    total_today,
        'allowed_today':  allowed_today,
        'denied_today':   denied_today,
        'by_type':        by_type,
        'date_filter':    date_filter,
        'type_filter':    type_filter,
        'status_filter':  status_filter,
    })



@login_required
@staff_member_required
def admin_notifications(request):
    if request.method == 'POST' and request.POST.get('action') == 'mark_all_read':
        Notification.objects.filter(user=request.user).update(is_read=True)
        return redirect('admin_notifications')
    notifs = Notification.objects.filter(user=request.user).order_by('-created_at')
    unread_count = notifs.filter(is_read=False).count()
    return render(request, 'attendance/admin_notifications.html', {
        'notifications': notifs, 'unread_count': unread_count,
    })


@login_required
def export_coordinator_students_csv(request):
    coordinator = get_object_or_404(Coordinator, auth_user=request.user)
    students = Student.objects.filter(
        department__college=coordinator.college).select_related('department')
    response = HttpResponse(content_type='text/csv; charset=utf-8-sig')
    from datetime import date
    college_name = coordinator.college.college_name.replace(' ', '_')[:25] if coordinator.college else 'College'
    response['Content-Disposition'] = f'attachment; filename="SHAMEL_{college_name}_Students_{date.today()}.csv"'
    writer = csv.writer(response)
    writer.writerow(['Student ID', 'Name', 'Batch', 'Registered', 'Phone', 'Email'])
    for s in students:
        writer.writerow([
            s.student_code, s.name, s.batch or '', s.is_registered,
            s.phone_number or '', s.university_email or '',
        ])
    return response


@login_required
def notifications_view(request):
    if request.method == 'POST':
        action   = request.POST.get('action', '')
        notif_id = request.POST.get('notif_id', '')
        if action == 'mark_read' and notif_id:
            Notification.objects.filter(id=notif_id, user=request.user).update(is_read=True)
        elif action == 'mark_all_read':
            Notification.objects.filter(user=request.user, is_read=False).update(is_read=True)
        elif action == 'delete' and notif_id:
            Notification.objects.filter(id=notif_id, user=request.user).delete()
        elif action == 'delete_all':
            Notification.objects.filter(user=request.user).delete()
        return redirect('notifications_view')

    notifs = Notification.objects.filter(user=request.user).order_by('-created_at')
    unread = notifs.filter(is_read=False).count()
    return render(request, 'attendance/notifications.html', {
        'notifications': notifs, 'unread': unread,
    })


@login_required
def mark_notification_read(request):
    if request.method == 'POST':
        notif_id = request.POST.get('notif_id', '')
        link = ''
        if notif_id:
            notif = Notification.objects.filter(id=notif_id, user=request.user).first()
            if notif:
                link = getattr(notif, 'link', '') or ''
                notif.is_read = True
                notif.save(update_fields=['is_read'])
        if link:
            return redirect(link)
        return redirect('notifications_view')
    return redirect('notifications_view')


class _FakeProfile:
    """Lightweight stand-in for a missing UserProfile model."""
    show_phone_to_peers = True
    show_email_to_peers = True
    show_attendance_to_coordinator = True
    email_notifications = True
    attendance_alerts = True
    ticket_updates = True
    weekly_summary = False
    require_face_login = False
    last_password_change = None


@login_required
def settings_view(request):
    # Use the real profile if it exists, otherwise fall back to the stub.
    try:
        from .models import UserProfile
        profile, _ = UserProfile.objects.get_or_create(user=request.user)
    except Exception:
        profile = _FakeProfile()

    student_obj  = Student.objects.filter(auth_user=request.user).first()
    teacher_obj  = Teacher.objects.filter(auth_user=request.user).first()
    return render(request, 'attendance/settings.html', {
        'profile':         profile,
        'user':            request.user,
        'student_profile': student_obj,
        'teacher_profile': teacher_obj,
    })


@login_required
def update_settings(request):
    if request.method != 'POST':
        return redirect('settings_page')

    tab = request.POST.get('tab', 'account')
    user = request.user

    if tab == 'account':
        user.first_name = request.POST.get('first_name', user.first_name).strip()
        user.last_name  = request.POST.get('last_name',  user.last_name).strip()
        user.email      = request.POST.get('email',      user.email).strip()
        user.save(update_fields=['first_name', 'last_name', 'email'])
        # Update face image if uploaded
        if 'profile_photo' in request.FILES:
            img = request.FILES['profile_photo']
            student = Student.objects.filter(auth_user=user).first()
            teacher = Teacher.objects.filter(auth_user=user).first()
            if student:
                student.face_image = img
                student.save(update_fields=['face_image'])
            elif teacher:
                teacher.face_image = img
                teacher.save(update_fields=['face_image'])
        messages.success(request, 'تم تحديث بيانات الحساب.')

    elif tab == 'security':
        current_pw  = request.POST.get('current_password', '')
        new_pw      = request.POST.get('new_password', '')
        confirm_pw  = request.POST.get('confirm_password', '')
        if new_pw:
            if not user.check_password(current_pw):
                messages.error(request, 'كلمة المرور الحالية غير صحيحة.')
            elif new_pw != confirm_pw:
                messages.error(request, 'كلمتا المرور غير متطابقتين.')
            elif len(new_pw) < 8:
                messages.error(request, 'يجب أن تكون كلمة المرور 8 أحرف على الأقل.')
            else:
                user.set_password(new_pw)
                user.save()
                from django.contrib.auth import update_session_auth_hash
                update_session_auth_hash(request, user)
                messages.success(request, 'تم تغيير كلمة المرور بنجاح.')
        else:
            messages.info(request, 'لم يتم تغيير كلمة المرور.')

    elif tab == 'privacy':
        try:
            from .models import UserProfile
            profile, _ = UserProfile.objects.get_or_create(user=user)
            profile.show_phone_to_peers            = 'show_phone' in request.POST
            profile.show_email_to_peers            = 'show_email' in request.POST
            profile.show_attendance_to_coordinator = 'show_attendance' in request.POST
            profile.save(update_fields=[
                'show_phone_to_peers', 'show_email_to_peers', 'show_attendance_to_coordinator'
            ])
            messages.success(request, 'تم حفظ إعدادات الخصوصية.')
        except Exception as e:
            messages.error(request, f'خطأ: {e}')

    elif tab == 'notifications':
        try:
            from .models import UserProfile
            profile, _ = UserProfile.objects.get_or_create(user=user)
            profile.email_notifications = 'email_notif' in request.POST
            profile.attendance_alerts   = 'att_alerts' in request.POST
            profile.ticket_updates      = 'ticket_updates' in request.POST
            profile.weekly_summary      = 'weekly_summary' in request.POST
            profile.save(update_fields=[
                'email_notifications', 'attendance_alerts', 'ticket_updates', 'weekly_summary'
            ])
            messages.success(request, 'تم حفظ إعدادات الإشعارات.')
        except Exception as e:
            messages.error(request, f'خطأ: {e}')

    return redirect('settings_page')


# ── Timetable Calendar ────────────────────────────────────────────────────────

@login_required
def schedule_calendar(request):
    """
    Timetable calendar view. Coordinator or admin only.
    GET params: sem_start, sem_end, semester
    Shows a 7-column week grid with existing Schedule slots.
    """
    coordinator = Coordinator.objects.filter(auth_user=request.user).first()
    is_admin    = request.user.is_staff or request.user.is_superuser
    if not (is_admin or coordinator):
        messages.error(request, 'غير مصرح.')
        return redirect('schedule')

    if coordinator and not is_admin:
        courses    = Course.objects.filter(college=coordinator.college).order_by('title')
        teachers   = Teacher.objects.filter(college=coordinator.college).order_by('name')
        classrooms = Classroom.objects.filter(
            Q(college=coordinator.college) | Q(college__isnull=True)
        ).order_by('name')
        base_schedules = Schedule.objects.filter(
            course__college=coordinator.college
        ).select_related('course', 'teacher', 'classroom')
    else:
        courses    = Course.objects.order_by('title')
        teachers   = Teacher.objects.order_by('name')
        classrooms = Classroom.objects.order_by('name')
        base_schedules = Schedule.objects.select_related('course', 'teacher', 'classroom').all()

    sem_start = request.GET.get('sem_start', '')
    sem_end   = request.GET.get('sem_end', '')
    semester  = request.GET.get('semester', '')

    if semester:
        base_schedules = base_schedules.filter(semester=semester)

    DAYS_ORDER = ['Saturday', 'Sunday', 'Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday']
    DAY_AR = {
        'Saturday': 'السبت', 'Sunday': 'الأحد', 'Monday': 'الاثنين',
        'Tuesday': 'الثلاثاء', 'Wednesday': 'الأربعاء',
        'Thursday': 'الخميس', 'Friday': 'الجمعة',
    }
    schedules_by_day = {d: [] for d in DAYS_ORDER}
    for s in base_schedules:
        if s.day_of_week in schedules_by_day:
            schedules_by_day[s.day_of_week].append(s)

    days_data = [
        {'key': d, 'ar': DAY_AR[d], 'slots': schedules_by_day[d]}
        for d in DAYS_ORDER
    ]

    return render(request, 'attendance/timetable_calendar.html', {
        'days_data':   days_data,
        'courses':     courses,
        'teachers':    teachers,
        'classrooms':  classrooms,
        'sem_start':   sem_start,
        'sem_end':     sem_end,
        'semester':    semester,
        'semesters':   SEMESTER_CHOICES_4Y,
        'coordinator': coordinator,
        'is_admin':    is_admin,
    })


@login_required
@require_POST
def calendar_add_slot(request):
    """
    Add one or more Schedule entries from the calendar drawer.
    Supports multiple teachers via teacher_id[] list.
    """
    coordinator = Coordinator.objects.filter(auth_user=request.user).first()
    is_admin    = request.user.is_staff or request.user.is_superuser
    if not (is_admin or coordinator):
        return JsonResponse({'ok': False, 'error': 'غير مصرح'}, status=403)

    course_id    = request.POST.get('course_id')
    day_of_week  = request.POST.get('day_of_week')
    start_time   = request.POST.get('start_time')
    end_time     = request.POST.get('end_time')
    semester     = request.POST.get('semester', '')
    classroom_id = request.POST.get('classroom_id') or None
    teacher_ids  = request.POST.getlist('teacher_id[]')
    if not teacher_ids:
        single = request.POST.get('teacher_id')
        teacher_ids = [single] if single else []

    if not course_id or not day_of_week or not start_time or not end_time:
        messages.error(request, 'بيانات ناقصة.')
        return redirect('schedule_calendar')

    if coordinator and not is_admin:
        if not Course.objects.filter(pk=course_id, college=coordinator.college).exists():
            messages.error(request, 'المادة ليست في كليتك.')
            return redirect('schedule_calendar')

    from .email_utils import notify_teacher_assignment
    created_count = 0
    if teacher_ids:
        for tid in teacher_ids:
            if not tid:
                continue
            try:
                sched = Schedule.objects.create(
                    course_id=course_id, teacher_id=tid,
                    classroom_id=classroom_id, day_of_week=day_of_week,
                    start_time=start_time, end_time=end_time, semester=semester,
                )
                created_count += 1
                try:
                    teacher = Teacher.objects.get(pk=tid)
                    notify_teacher_assignment(teacher, sched)
                except Exception:
                    pass
            except Exception as e:
                messages.error(request, f'خطأ: {e}')
    if not created_count:
        try:
            Schedule.objects.create(
                course_id=course_id, teacher_id=None,
                classroom_id=classroom_id, day_of_week=day_of_week,
                start_time=start_time, end_time=end_time, semester=semester,
            )
        except Exception as e:
            messages.error(request, f'خطأ: {e}')

    messages.success(request, 'تمت إضافة الحصة الدراسية.')
    sem_start = request.POST.get('sem_start', '')
    sem_end   = request.POST.get('sem_end', '')
    return redirect(f"/schedule/calendar/?sem_start={sem_start}&sem_end={sem_end}&semester={semester}")


@login_required
def calendar_delete_slot(request, schedule_id):
    """Delete a schedule slot from the calendar."""
    coordinator = Coordinator.objects.filter(auth_user=request.user).first()
    is_admin    = request.user.is_staff or request.user.is_superuser
    if not (is_admin or coordinator):
        messages.error(request, 'غير مصرح.')
        return redirect('schedule_calendar')

    schedule = get_object_or_404(Schedule, pk=schedule_id)
    if coordinator and not is_admin:
        if schedule.course.college != coordinator.college:
            messages.error(request, 'لا يمكنك حذف جداول كليات أخرى.')
            return redirect('schedule_calendar')
    schedule.delete()
    messages.success(request, 'تم حذف الحصة.')
    sem = request.GET.get('semester', '')
    ss  = request.GET.get('sem_start', '')
    se  = request.GET.get('sem_end', '')
    return redirect(f"/schedule/calendar/?sem_start={ss}&sem_end={se}&semester={sem}")


# ── Attendance Warning Check ──────────────────────────────────────────────────

@login_required
def check_attendance_warnings(request):
    """
    Staff-only view: scan all students and send warning emails
    when attendance is below 80%.
    """
    if not (request.user.is_staff or request.user.is_superuser):
        return JsonResponse({'error': 'غير مصرح'}, status=403)

    from .email_utils import notify_student_attendance_warning
    sent = 0
    students = Student.objects.all()
    for student in students:
        logs = AIAttendanceLog.objects.filter(student=student)
        total = logs.count()
        if total == 0:
            continue
        present = logs.filter(status='Present').count()
        pct = present / total * 100
        if pct < 80:
            notify_student_attendance_warning(student, pct)
            sent += 1

    return JsonResponse({'ok': True, 'emails_sent': sent})


# ── PWA Views ─────────────────────────────────────────────────────────────────

from django.views.decorators.cache import cache_control
from django.views.decorators.http import require_GET
import json as _json

@require_GET
@cache_control(max_age=0, no_cache=True, no_store=True, must_revalidate=True)
def pwa_sw(request):
    """Serve the service worker with a dynamic version injected (git hash or timestamp)."""
    import os, subprocess
    from django.conf import settings

    # Dynamic version = git short hash; fall back to the sw.js file mtime so the
    # version bumps whenever the worker changes even without git. This forces the
    # browser to install the new SW (which purges old caches) on every change.
    import re as _re
    sw_path = os.path.join(settings.BASE_DIR, 'attendance', 'static', 'pwa', 'sw.js')
    try:
        version = subprocess.check_output(
            ['git', 'rev-parse', '--short', 'HEAD'],
            cwd=settings.BASE_DIR, stderr=subprocess.DEVNULL
        ).decode().strip()
    except Exception:
        version = ''
    try:
        version = f"{version}-{int(os.path.getmtime(sw_path))}"
    except Exception:
        from datetime import datetime
        version = version or datetime.now().strftime('%Y%m%d%H%M%S')

    with open(sw_path, 'r', encoding='utf-8') as f:
        content = f.read()

    # Robustly replace whatever the hard-coded VERSION line is with the live one
    # (regex tolerates any spacing / current version string).
    content = _re.sub(r"const\s+VERSION\s*=\s*'[^']*';",
                      f"const VERSION = 'shamel-{version}';", content, count=1)

    resp = HttpResponse(content, content_type='application/javascript; charset=utf-8')
    resp['Cache-Control'] = 'no-cache, no-store, must-revalidate'
    return resp


@require_GET
def pwa_offline(request):
    """Offline fallback page."""
    return render(request, 'attendance/offline.html')


@require_GET
def pwa_manifest(request):
    """Serve manifest.json from root."""
    import os
    from django.conf import settings
    manifest_path = os.path.join(settings.BASE_DIR, 'attendance', 'static', 'pwa', 'manifest.json')
    with open(manifest_path, 'r', encoding='utf-8') as f:
        data = _json.load(f)
    return JsonResponse(data, json_dumps_params={'ensure_ascii': False, 'indent': 2})


# ── Live Reload Trigger ───────────────────────────────────────────────────────

@require_GET
def trigger_live_reload(request):
    """
    Staff-only endpoint: broadcast a reload signal to all connected browsers.
    Called automatically by the deploy script after collectstatic.
    GET /api/live-reload/?secret=DEPLOY_SECRET
    """
    from django.conf import settings
    secret = request.GET.get('secret', '')
    deploy_secret = getattr(settings, 'DEPLOY_SECRET', '')

    # Must be staff or provide the deploy secret
    authed = (request.user.is_authenticated and request.user.is_staff) or \
             (deploy_secret and secret == deploy_secret)
    if not authed:
        return JsonResponse({'error': 'غير مصرح'}, status=403)

    import subprocess
    from asgiref.sync import async_to_sync
    from channels.layers import get_channel_layer

    try:
        version = subprocess.check_output(
            ['git', 'rev-parse', '--short', 'HEAD'],
            cwd=settings.BASE_DIR, stderr=subprocess.DEVNULL
        ).decode().strip()
    except Exception:
        version = 'unknown'

    channel_layer = get_channel_layer()
    async_to_sync(channel_layer.group_send)(
        'shamel_live_reload',
        {
            'type':    'live_reload',
            'reason':  'deploy',
            'version': version,
        }
    )
    return JsonResponse({'ok': True, 'version': version, 'message': 'Reload signal sent'})


def api_ping(request):
    """Lightweight connectivity check — no auth required.
    The PWA uses this instead of navigator.onLine so that devices on a
    local/emulator network (with server reachable but no internet) are
    correctly treated as online.
    """
    from django.http import JsonResponse
    return JsonResponse({'ok': True, 'status': 'online'})


def api_list_cameras(request):
    """Return list of available camera indices on the server.
    Used by the scan page to let users select an external USB camera."""
    if not CV2_AVAILABLE:
        return JsonResponse({'cameras': [{'index': 0, 'label': 'Default Camera'}]})
    cameras = []
    for i in range(5):  # check indices 0-4
        try:
            cap = cv2.VideoCapture(i, cv2.CAP_DSHOW)
            if cap.isOpened():
                ret, _ = cap.read()
                cap.release()
                if ret:
                    cameras.append({'index': i, 'label': f'Camera {i}' if i > 0 else 'Built-in Camera (0)'})
        except Exception:
            pass
    if not cameras:
        cameras = [{'index': 0, 'label': 'Default Camera'}]
    return JsonResponse({'cameras': cameras})
